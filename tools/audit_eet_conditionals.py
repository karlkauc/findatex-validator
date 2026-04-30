"""Surface candidate conditional rules in the EET spec by regex-scanning
the COMMENT column for "conditional to / if (item|field|number) /
when ... set to" phrases. Filters out NUMs already wired in
``EetRuleSet.CONDITIONAL_REQUIREMENTS`` plus the three new rules added by
this branch (negative SFDR, PAI gating, taxonomy at-least-one).

Triage-only — does not modify code. Pure read-only.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

from findatex_sample_helpers import ROOT, load_spec

# NUMs already wired (existing 6 + 3 new from this branch). Keys are
# target NUMs; the audit reports candidates whose NUM is NOT in this set.
WIRED_TARGETS = {
    # existing 6
    "28", "30", "31", "41", "45", "48",
    # 3 new — negative-SFDR forbidden Art-fields, PAI-block, taxonomy at-least-one
    "40", "42", "43", "44", "46", "47",
    "103", "104", "106", "110", "114", "118", "122", "126", "130",
    "134", "138", "142", "146", "150", "154", "158", "162", "166",
    "170", "174", "178", "182", "186", "190", "194", "198", "202",
}

PHRASE_RE = re.compile(
    r"(?i)conditional to|if (?:item|field|number)|"
    r"when .* set to|set to \"?[ynYN0-9]+\"?",
)


def comments_for(version: str) -> dict[str, str]:
    manifest_rel, xlsx_rel = {
        "V1.1.2": (
            "core/src/main/resources/spec/eet/eet-v112-info.json",
            "core/src/main/resources/spec/eet/EET_V1_1_2_20231205.xlsx",
        ),
        "V1.1.3": (
            "core/src/main/resources/spec/eet/eet-v113-info.json",
            "core/src/main/resources/spec/eet/EET_V1_1_3_20260410.xlsx",
        ),
    }[version]
    manifest = json.loads((ROOT / manifest_rel).read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(ROOT / xlsx_rel, data_only=True)
    ws = wb[manifest["sheetName"]]
    out: dict[str, str] = {}
    for r in range(manifest["firstDataRow"], ws.max_row + 1):
        n = ws.cell(r, manifest["columns"]["numData"]).value
        if n is None:
            continue
        out[str(n).strip()] = str(ws.cell(r, manifest["columns"]["comment"]).value or "")
    return out


def main() -> int:
    for version in ("V1.1.2", "V1.1.3"):
        _, fields = load_spec("EET", version)
        comments = comments_for(version)
        print(f"## EET {version} — conditional candidates not yet wired")
        print()
        print("| NUM | path | comment excerpt |")
        print("|----:|:-----|:----------------|")
        any_row = False
        for f in fields:
            if f.num in WIRED_TARGETS:
                continue
            comment = comments.get(f.num, "")
            if not PHRASE_RE.search(comment):
                continue
            short = comment.replace("\n", " ⏎ ").strip()
            if len(short) > 100:
                short = short[:97] + "..."
            print(f"| {f.num} | {f.path} | {short} |")
            any_row = True
        if not any_row:
            print("| _(none)_ | | |")
        print()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
