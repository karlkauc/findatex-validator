"""Per-profile M-flag diff between EET V1.1.2 and V1.1.3.

Walks every profile column declared in each manifest, reads the M/C/O
flag per row, and reports symmetric differences (gained-M, lost-M,
demoted-to-C, promoted-from-C). Pure read-only.
"""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

import openpyxl

from findatex_sample_helpers import ROOT


def load(version: str) -> dict[str, dict[str, str]]:
    manifest_rel, xlsx_rel = {
        "V1.1.2": (
            "core/src/main/resources/spec/eet/eet-v112-info.json",
            "core/src/main/resources/spec/eet/EET_V1_1_2_20231205.xlsx",
        ),
        "V1.1.3": (
            "core/src/main/resources/spec/eet/eet-v113-info.json",
            "core/src/main/resources/spec/eet/EET_V1_1_3_20260410.xlsx",
        ),
    }[version]
    manifest = json.loads((ROOT / manifest_rel).read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(ROOT / xlsx_rel, data_only=True)
    ws = wb[manifest["sheetName"]]

    # profile -> { num -> flag }
    out: dict[str, dict[str, str]] = defaultdict(dict)
    for r in range(manifest["firstDataRow"], ws.max_row + 1):
        num = ws.cell(r, manifest["columns"]["numData"]).value
        path = ws.cell(r, manifest["columns"]["path"]).value
        if num is None or path is None:
            continue
        n = str(num).strip()
        for p in manifest["profileColumns"]:
            v = ws.cell(r, p["column"]).value
            if v:
                out[p["code"]][n] = str(v).strip().upper()
    return dict(out)


def main() -> int:
    v112 = load("V1.1.2")
    v113 = load("V1.1.3")

    profiles = sorted(set(v112) | set(v113))
    print("## M-flag drift V1.1.2 → V1.1.3")
    print()
    print("| profile | gained M | lost M | other change |")
    print("|:--------|:---------|:-------|:-------------|")
    for prof in profiles:
        a = v112.get(prof, {})
        b = v113.get(prof, {})
        nums = sorted(set(a) | set(b), key=lambda s: int(s) if s.isdigit() else 9999)

        gained_m = [n for n in nums if a.get(n) != "M" and b.get(n) == "M"]
        lost_m = [n for n in nums if a.get(n) == "M" and b.get(n) != "M"]
        other = [
            f"{n}: {a.get(n) or '-'}→{b.get(n) or '-'}"
            for n in nums
            if a.get(n) != b.get(n)
            and not (a.get(n) != "M" and b.get(n) == "M")
            and not (a.get(n) == "M" and b.get(n) != "M")
        ]

        def fmt(xs: list[str]) -> str:
            if not xs:
                return "—"
            short = ", ".join(xs[:8])
            if len(xs) > 8:
                short += f" (+{len(xs) - 8} more)"
            return short

        print(f"| {prof} | {fmt(gained_m)} | {fmt(lost_m)} | {fmt(other)} |")
    print()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
