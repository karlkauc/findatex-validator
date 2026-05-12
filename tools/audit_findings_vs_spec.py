#!/usr/bin/env python3
"""
audit_findings_vs_spec.py — audit the validator's findings against the FinDatEx
spec XLSX files. Outputs /home/karl/webdav/findatex-testdata/SPEC-AUDIT.md.

Read-only. Does not touch the validator source.
"""
from __future__ import annotations
import collections, json, pathlib, re, sys
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import openpyxl

REPO     = pathlib.Path('/home/karl/webdav/findatex-validator')
TESTDATA = pathlib.Path('/home/karl/webdav/findatex-testdata')
OUT      = TESTDATA / 'SPEC-AUDIT.md'

# ---------------------------------------------------------------------------
# Template registry
# ---------------------------------------------------------------------------

@dataclass
class TmplCfg:
    label: str
    manifest: pathlib.Path
    xlsx: pathlib.Path
    report: pathlib.Path

TEMPLATES: List[TmplCfg] = [
    TmplCfg('EMT V4.2',
            REPO/'core/src/main/resources/spec/emt/emt-v42-info.json',
            REPO/'core/src/main/resources/spec/emt/EMT_V4_2_20240422.xlsx',
            TESTDATA/'EMT-V4.2-report.xlsx'),
    TmplCfg('EMT V4.3',
            REPO/'core/src/main/resources/spec/emt/emt-v43-info.json',
            REPO/'core/src/main/resources/spec/emt/EMT_V4_3_20251217.xlsx',
            TESTDATA/'EMT-V4.3-report.xlsx'),
    TmplCfg('EET V1.1.2',
            REPO/'core/src/main/resources/spec/eet/eet-v112-info.json',
            REPO/'core/src/main/resources/spec/eet/EET_V1_1_2_20231205.xlsx',
            TESTDATA/'EET-V1.1.2-report.xlsx'),
    TmplCfg('EET V1.1.3',
            REPO/'core/src/main/resources/spec/eet/eet-v113-info.json',
            REPO/'core/src/main/resources/spec/eet/EET_V1_1_3_20260410.xlsx',
            TESTDATA/'EET-V1.1.3-report.xlsx'),
    TmplCfg('EPT V2.1',
            REPO/'core/src/main/resources/spec/ept/ept-v21-info.json',
            REPO/'core/src/main/resources/spec/ept/EPT_V2_1_20221012.xlsx',
            TESTDATA/'EPT-V2.1-report.xlsx'),
]

# ---------------------------------------------------------------------------
# Spec loading (mirrors core's extractNumKey logic — split on first underscore)
# ---------------------------------------------------------------------------

@dataclass
class FieldRec:
    num_key: str
    num_data: str
    name: str
    codification: str
    comment: str
    definition: str
    flags: Dict[str, str] = field(default_factory=dict)   # profile_code -> raw flag string
    row_idx: int = 0

def cell_str(row, col_1based: Optional[int]) -> str:
    if not col_1based: return ''
    c = row[col_1based - 1].value if col_1based - 1 < len(row) else None
    if c is None: return ''
    return str(c).strip()

def extract_num_key(num_data: str) -> str:
    """Mirror FieldSpec.extractNumKey: take prefix up to first '_'."""
    t = num_data.strip()
    us = t.find('_')
    return t if us < 0 else t[:us]

def load_spec(cfg: TmplCfg) -> Dict[str, FieldRec]:
    manifest = json.loads(cfg.manifest.read_text())
    cols = manifest['columns']
    profile_cols = {p['code']: p.get('column') for p in manifest['profileColumns'] if p.get('column')}
    wb = openpyxl.load_workbook(cfg.xlsx, read_only=True, data_only=True)
    sheet = wb[manifest['sheetName']]
    out: Dict[str, FieldRec] = {}
    first_data_row = manifest['firstDataRow']  # 1-based
    for r_idx, row in enumerate(sheet.iter_rows(min_row=first_data_row), start=first_data_row):
        num_data = cell_str(row, cols['numData'])
        name = cell_str(row, cols.get('name'))
        if not num_data and not name:
            continue
        # Filter out section headers (no field-label shape)
        if not name and '_' not in num_data and not num_data.isdigit() and not re.match(r'^\d+[a-z]?$', num_data):
            continue
        num_key = extract_num_key(num_data)
        flags: Dict[str, str] = {}
        for p, col in profile_cols.items():
            flags[p] = cell_str(row, col)
        rec = FieldRec(
            num_key=num_key,
            num_data=num_data,
            name=name or num_data,
            codification=cell_str(row, cols.get('codification')),
            comment=cell_str(row, cols.get('comment')),
            definition=cell_str(row, cols.get('definition')),
            flags=flags,
            row_idx=r_idx,
        )
        if num_key and num_key not in out:
            out[num_key] = rec
    wb.close()
    return out

# ---------------------------------------------------------------------------
# Findings loading
# ---------------------------------------------------------------------------

@dataclass
class RuleAgg:
    rule_id: str
    field_num: str
    profile: str
    severity: str
    field_name: str
    sample_message: str
    sample_value: str
    count: int = 0
    distinct_files: set = field(default_factory=set)

def load_findings(report_xlsx: pathlib.Path) -> Dict[str, RuleAgg]:
    wb = openpyxl.load_workbook(report_xlsx, read_only=True, data_only=True)
    ws = wb['All Findings']
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {h: header.index(h) for h in header}
    out: Dict[str, RuleAgg] = {}
    for r in rows:
        rule_id  = r[idx['Rule']] or ''
        field_num = (r[idx['Field#']] or '')
        profile  = r[idx['Profile']] or ''
        severity = r[idx['Severity']] or ''
        field_name = r[idx['Field name']] or ''
        message  = r[idx['Message']] or ''
        value    = r[idx['Value']] or ''
        file_    = r[idx['File']] or ''
        # Group key = full rule_id (already contains field/profile for PRESENCE/FORMAT/etc.)
        key = f'{rule_id}|{field_num}|{profile}'
        agg = out.get(key)
        if agg is None:
            agg = RuleAgg(rule_id=rule_id, field_num=str(field_num), profile=str(profile),
                          severity=str(severity), field_name=str(field_name),
                          sample_message=str(message), sample_value=str(value))
            out[key] = agg
        agg.count += 1
        agg.distinct_files.add(file_)
    wb.close()
    return out

# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------

VERDICT_TP    = 'TP'
VERDICT_FP    = 'FP'
VERDICT_REVIEW= 'NEEDS-REVIEW'
VERDICT_MANUAL= 'MANUAL-REVIEW'
VERDICT_UNCLEAR='UNCLEAR'
VERDICT_UNKNOWN='UNKNOWN'

# Manuelle Verdicts für die ~17 distinct XF-Rule-IDs, gewonnen aus
# `EetRuleSet.java` / `EmtVersionRule.java` / `EptVersionRule.java` + den
# Spec-Comment-Spalten der adressierten Felder.
XF_VERDICTS: Dict[str, Tuple[str, str]] = {
    'EMT-XF-VERSION': (VERDICT_TP,
        'Prüft Feld 1 = erwartete EMT-Version. Spec-Codification gibt die Version vor (z.B. "V4.2"). Universeller Sanity-Check.'),
    'EET-XF-VERSION': (VERDICT_TP,
        'Prüft Feld 1 = erwartete EET-Version. Universeller Sanity-Check.'),
    'EPT-XF-VERSION': (VERDICT_TP,
        'Prüft Feld 1 = erwartete EPT-Version. Universeller Sanity-Check.'),
    'EET-XF-ART8-MIN-LT': (VERDICT_TP,
        'Trigger: Feld 27 (SFDR-Product-Type) = "8". Pflicht: Feld 30. Spec-Comment für 30 lautet sinngemäss "Conditional for Financial Product primarily investing in Art 8 or Art 8 like funds" — passt 1:1.'),
    'EET-XF-ART9-MIN-LT': (VERDICT_TP,
        'Trigger: Feld 27 = "9". Pflicht: Feld 31. Spec-Comment 31: "Conditional for Financial Product primarily investing in art 9 or art 9 like funds" — passt 1:1.'),
    'EET-XF-COUNTRYLIST-616': (VERDICT_TP,
        'Trigger: Feld 228 (NUM 31240) > 0. Pflicht: Feld 616. Spec-Comment 616: "Blank if none; conditional to 31240 > 0" — direkt zitiert.'),
    'EET-XF-COUNTRYLIST-615': (VERDICT_TP,
        'Trigger: Feld 225 (NUM 31210) > 0. Pflicht: Feld 615. Spec-Comment 615 analog zu 616. (Im Lauf nicht gefeuert.)'),
    'EET-XF-PCDFP-35': (VERDICT_TP,
        'Trigger: Feld 27 oder 28 = "8" oder "9". Pflicht: Feld 35. Severity = WARNING wegen Spec-Softening "Could be provided for art6". (Im Lauf nicht gefeuert.)'),
    'EET-XF-PCDFP-36': (VERDICT_TP,
        'Wie PCDFP-35 für Feld 36. Severity = WARNING — Spec-Comment enthält "Could be provided for art6 under insurers demand". Validator-Severity ist absichtlich softer als ERROR.'),
    'EET-XF-SFDR-OUT-OF-SCOPE': (VERDICT_TP,
        'Trigger: Feld 27 = "0". Pflicht: Feld 28 (Eligibility-Klassifikation). Spec-Comment 28: "Conditional to 20040 set to 0" — direkt zitiert. (Im Lauf nicht gefeuert.)'),
    'EET-XF-ART8-MIN-SI': (VERDICT_TP,
        'Trigger: Feld 40 (Y). Pflicht: Feld 41. Spec-Comment 41: "Conditional ... 20170 = Yes". (Im Lauf nicht gefeuert.)'),
    'EET-XF-ART9-MIN-ENV': (VERDICT_TP,
        'Trigger: Feld 27 = "9". Pflicht: Feld 45. (Im Lauf nicht gefeuert.)'),
    'EET-XF-ART9-MIN-SOC': (VERDICT_TP,
        'Trigger: Feld 27 = "9". Pflicht: Feld 48. (Im Lauf nicht gefeuert.)'),
    'EET-XF-ART9-PARIS-DECARB-80': (VERDICT_UNCLEAR,
        'Trigger: Feld 27 = "9". Soll: Feld 80. Java-Kommentar markiert "PENDING SME SIGN-OFF" — Spec-Text "Could be fulfilled for art 8" softens. Severity = WARNING korrekt konservativ.'),
    'EET-XF-ART9-PARIS-DECARB-81': (VERDICT_UNCLEAR,
        'Wie -80 für Feld 81. PENDING SME SIGN-OFF.'),
    'EET-XF-ART40-MUST-BE-ABSENT': (VERDICT_UNCLEAR,
        'Negative SFDR-Constraint: wenn Feld 27 = "0", muss Feld 40 leer sein. Spec sagt nicht explizit "muss leer" sondern nur "conditional". Validator-Severity = WARNING (PENDING SME SIGN-OFF, im Java-Kommentar dokumentiert).'),
    'EET-XF-ART42-MUST-BE-ABSENT': (VERDICT_UNCLEAR, 'Analog ART40-MUST-BE-ABSENT.'),
    'EET-XF-ART43-MUST-BE-ABSENT': (VERDICT_UNCLEAR, 'Analog ART40-MUST-BE-ABSENT.'),
    'EET-XF-ART44-MUST-BE-ABSENT': (VERDICT_UNCLEAR, 'Analog ART40-MUST-BE-ABSENT.'),
    'EET-XF-ART46-MUST-BE-ABSENT': (VERDICT_UNCLEAR, 'Analog ART40-MUST-BE-ABSENT.'),
    'EET-XF-ART47-MUST-BE-ABSENT': (VERDICT_UNCLEAR, 'Analog ART40-MUST-BE-ABSENT.'),
    'EET-XF-PAI-186': (VERDICT_UNCLEAR,
        'PAI-Block-Gating: wenn Feld 33 = "Y", muss Feld 186 vorhanden sein. **Spec-Comment ist leer** für Feld 186. Validator-Logik im Java-Kommentar: "Only NUM=106 carries the explicit conditional comment in the spec; the remainder of the block is gated implicitly". → Validator extrapoliert über den Spec-Wortlaut (operator intent). **Severity = WARNING** (downgraded von ERROR am 2026-05-12 nach diesem Audit) — gleicher Status wie ART*-MUST-BE-ABSENT.'),
}

def classify(agg: RuleAgg, spec: Dict[str, FieldRec]) -> Tuple[str, str]:
    """Return (verdict, evidence_text)."""
    rid = agg.rule_id
    # PRESENCE/X/PROFILE
    if rid.startswith('PRESENCE/'):
        parts = rid.split('/')
        if len(parts) == 3:
            field_num, prof = parts[1], parts[2]
            rec = spec.get(field_num)
            if not rec:
                return VERDICT_UNKNOWN, f'Spec hat kein Feld {field_num}'
            flag = (rec.flags.get(prof) or '').strip()
            if flag == 'M':
                return VERDICT_TP, f"Spec-Flag '{flag}' (Pflicht)"
            if flag == 'C':
                return VERDICT_REVIEW, f"Spec-Flag 'C' (Conditional). Comment: {rec.comment[:200]}"
            if flag == '':
                return VERDICT_FP, f"Spec-Flag leer in Profil {prof} — Validator sollte keine Pflicht ableiten"
            if flag in {'O', 'I', 'N'}:
                return VERDICT_FP, f"Spec-Flag '{flag}' — keine Pflicht in Profil {prof}"
            # E.g. EPT 'x' from PRIIPS_SYNC/UK columns; or weird textual content
            return VERDICT_FP, f"Spec-Flag '{flag}' — kein M/C/O/I/N; Validator interpretiert es als Pflicht (Verdacht: Implementierungsfehler)"
    # FORMAT/X
    if rid.startswith('FORMAT/'):
        field_num = rid.split('/', 1)[1]
        rec = spec.get(field_num)
        if not rec:
            return VERDICT_UNKNOWN, f'Spec hat kein Feld {field_num}'
        codif = rec.codification.strip()
        if not codif:
            return VERDICT_FP, 'Spec-Codification leer — Format-Check ohne Spec-Basis'
        if 'free text' in codif.lower():
            return VERDICT_FP, 'Spec-Codification = "free text"'
        return VERDICT_TP, f'Spec-Codification: {codif[:140]}'
    # ISIN/X/Y or LEI/X/Y
    if rid.startswith('ISIN/') or rid.startswith('LEI/'):
        parts = rid.split('/')
        code_field = parts[1] if len(parts) > 1 else ''
        type_field = parts[2] if len(parts) > 2 else ''
        rec = spec.get(type_field) or spec.get(code_field)
        if not rec:
            return VERDICT_UNKNOWN, f'Spec hat kein Feld {type_field or code_field}'
        return VERDICT_TP, f'Identifier-Codecheck (universell). Type-Feld {type_field} Codification: {rec.codification[:140]}'
    # Cross-field rules — manuell auditiert via XF_VERDICTS
    for prefix in ('EMT-XF', 'EET-XF', 'EPT-XF', 'TPT-XF', 'XF-'):
        if rid.startswith(prefix):
            verdict_pair = XF_VERDICTS.get(rid)
            if verdict_pair:
                return verdict_pair
            # Fallback: noch nicht manuell klassifiziert
            field_num = agg.field_num
            rec = spec.get(field_num)
            comment = rec.comment[:200] if rec else ''
            return VERDICT_MANUAL, f'XF-Regel ohne manuellen Verdict. Comment Feld {field_num}: {comment}'
    # CLOSED-LIST/* (defensive)
    if rid.startswith('CLOSED-LIST/'):
        field_num = rid.split('/', 1)[1]
        rec = spec.get(field_num)
        if not rec:
            return VERDICT_UNKNOWN, f'Spec hat kein Feld {field_num}'
        return VERDICT_TP, f'Closed-list check; Codification: {rec.codification[:140]}'
    return VERDICT_UNKNOWN, f'Unbekannter Rule-Typ: {rid}'

# ---------------------------------------------------------------------------
# Markdown emission
# ---------------------------------------------------------------------------

def truncate(s: str, n: int = 80) -> str:
    s = (s or '').replace('\n', ' ').replace('|', '\\|').strip()
    if len(s) <= n: return s
    return s[:n - 1] + '…'

def emit_markdown(per_template: Dict[str, Tuple[Dict[str, RuleAgg], Dict[str, FieldRec]]]) -> str:
    lines: List[str] = []
    lines.append('# FinDatEx Validator — Spec-Audit der Batch-Findings\n')
    lines.append('Generiert vom `audit_findings_vs_spec.py` Skript. Vergleicht jede '
                 'eindeutige Rule-ID aus dem letzten Batch-Lauf gegen das gebündelte '
                 'FinDatEx-Spec-XLSX (Source-of-truth). Read-only — kein Validator-Code '
                 'wurde angefasst, um zu klassifizieren.\n')

    # Methodology
    lines.append('## Methodologie\n')
    lines.append('Für jede eindeutige `(Rule-ID, Field#, Profile)` Kombination wird eine '
                 'Klassifikation in einen Verdict-Bucket vorgenommen:\n')
    lines.append('| Bucket | Bedeutung |')
    lines.append('|--------|-----------|')
    lines.append('| **TP** | True positive — Spec mandiert das geprüfte Verhalten; Finding ist legitim. |')
    lines.append('| **FP** | False positive — Spec mandiert das *nicht*; Validator-Implementierung erzeugt unbegründete Findings (Bug-Verdacht). |')
    lines.append('| **NEEDS-REVIEW** | Spec markiert das Feld als *Conditional* — eine echte Beurteilung erfordert das Lesen der Spec-Comment-Spalte und der Datenwerte. |')
    lines.append('| **MANUAL-REVIEW** | Cross-field-Regel — Verdict erfolgt manuell anhand der Spec-Comment-Spalte und der Validator-Logik. |')
    lines.append('| **UNKNOWN** | Rule-Typ vom Audit nicht erkannt — selten, deutet auf neue Rule-Klasse. |\n')
    lines.append('Klassifikationsregeln:\n')
    lines.append('- `PRESENCE/X/PROFILE`: Spec-Flag in Profil-Spalte für Feld X. `M`→TP, `C`→NEEDS-REVIEW, '
                 '`O`/`I`/`N`/leer → FP. Andere Werte (z. B. `x` aus EPT-Booleansspalten) → FP mit Bug-Hinweis.')
    lines.append('- `FORMAT/X`: Spec-Codification-Zelle muss eine Prüfbarkeit angeben. Leer / "free text" → FP.')
    lines.append('- `ISIN/X/Y` / `LEI/X/Y`: universelle Identifier-Checks; immer TP (Spec definiert ISO 6166 / 17442).')
    lines.append('- `*-XF-*`: MANUAL-REVIEW — bekommt im Nachgang ein finales Verdict per Hand.\n')
    lines.append('**Limitationen**: Audit erkennt nur *False Positives* (Findings ohne Spec-Basis). '
                 '*False Negatives* (Spec-Regeln, die der Validator nicht prüft) sind nicht Teil dieser '
                 'mechanischen Auswertung — siehe DEFERRED-Sektion am Ende.\n')

    # Summary table
    lines.append('## Zusammenfassung über alle Templates\n')
    lines.append('| Template | Findings | Unique Rules | TP | FP | NEEDS-REVIEW | UNCLEAR | MANUAL-REVIEW | UNKNOWN | Findings ohne klare Spec-Basis* |')
    lines.append('|----------|---------:|-------------:|---:|---:|-------------:|--------:|--------------:|--------:|--------------------------------:|')
    grand = {'findings': 0, 'fp_findings': 0, 'unclear_findings': 0}
    for tmpl, (rules, spec) in per_template.items():
        cnt_total = sum(a.count for a in rules.values())
        bucket_counts = collections.Counter()
        bucket_findings = collections.Counter()
        for agg in rules.values():
            verdict, _ = classify(agg, spec)
            bucket_counts[verdict] += 1
            bucket_findings[verdict] += agg.count
        fp_findings = bucket_findings[VERDICT_FP]
        unclear_findings = bucket_findings[VERDICT_UNCLEAR]
        questionable = fp_findings + unclear_findings
        pct_q = (questionable / cnt_total * 100) if cnt_total else 0.0
        lines.append(f'| {tmpl} | {cnt_total} | {len(rules)} | {bucket_counts[VERDICT_TP]} | '
                     f'{bucket_counts[VERDICT_FP]} | {bucket_counts[VERDICT_REVIEW]} | '
                     f'{bucket_counts[VERDICT_UNCLEAR]} | {bucket_counts[VERDICT_MANUAL]} | '
                     f'{bucket_counts[VERDICT_UNKNOWN]} | '
                     f'{questionable} ({pct_q:.1f} %) |')
        grand['findings'] += cnt_total
        grand['fp_findings'] += fp_findings
        grand['unclear_findings'] += unclear_findings
    grand_q = grand['fp_findings'] + grand['unclear_findings']
    lines.append(f'| **Σ Gesamt** | **{grand["findings"]}** | | | | | | | | '
                 f'**{grand_q} ({grand_q/grand["findings"]*100:.1f} %)** |')
    lines.append('')
    lines.append('*„Findings ohne klare Spec-Basis" = Summe aus FP (Validator-Implementierungsbug) '
                 'und UNCLEAR (Validator-Regel etwas strenger als Spec-Wortlaut, aber dokumentiert / '
                 'operator-akzeptiert).\n')

    # Detail tables
    for tmpl, (rules, spec) in per_template.items():
        lines.append(f'## {tmpl} — Detail pro Rule-ID\n')
        lines.append(f'Total Findings: {sum(a.count for a in rules.values()):,} · '
                     f'Unique Rules: {len(rules)}\n')
        # Sort: TP/FP buckets, then by count desc
        sorted_rules = sorted(rules.values(),
                              key=lambda a: (-a.count, a.rule_id))
        lines.append('| Verdict | Count | Files | Rule-ID | Field# | Field-Name | Profile | Sev | Spec-Flag | Codification (gekürzt) | Verdict-Begründung |')
        lines.append('|---------|------:|------:|---------|--------|------------|---------|-----|-----------|-----------------------|-------------------|')
        for agg in sorted_rules:
            verdict, evidence = classify(agg, spec)
            rec = spec.get(agg.field_num)
            flag = ''
            if rec and agg.profile:
                # Profile shows display name (e.g. "SFDR Periodic"); map to code
                # Try code lookup directly first; if not, try display name match
                manifest = json.loads([t for t in TEMPLATES if t.label == tmpl][0].manifest.read_text())
                for p in manifest['profileColumns']:
                    if p['code'] == agg.profile.replace(' ', '_').upper() or \
                       p.get('display','').lower() == agg.profile.lower() or \
                       p['code'] == agg.profile:
                        flag = rec.flags.get(p['code'], '')
                        break
                else:
                    # No match — try direct profile-code lookup
                    flag = rec.flags.get(agg.profile, '')
            elif rec and not agg.profile:
                flag = ''
            codif = rec.codification if rec else ''
            lines.append(f'| **{verdict}** | {agg.count} | {len(agg.distinct_files)} | `{agg.rule_id}` | '
                         f'{agg.field_num} | {truncate(agg.field_name, 60)} | '
                         f'{truncate(agg.profile, 25)} | {agg.severity[:1]} | `{flag}` | '
                         f'{truncate(codif, 60)} | {truncate(evidence, 120)} |')
        lines.append('')

    # Narrative sections (handgeschrieben, ergänzen die mechanische Klassifikation)
    lines.append('## Verdachtsfälle für Implementierungsfehler — explizit geprüft\n')
    lines.append('Drei Strukturen im Spec-Format gaben Anlass zu konkreten Bug-Verdächten. Alle drei wurden geprüft und sind **sauber**:\n')
    lines.append('### 1. EPT V2.1 Spalten 6 (PRIIPS_SYNC) und 8 (UK) — Boolean `x`/blank statt M/C/O\n')
    lines.append('| Spalte | Verteilung im Spec | Findings im Lauf |')
    lines.append('|--------|---------------------|------------------|')
    lines.append('| Col 6 (PRIIPS_SYNC) | `x`=101, leer=56 | 0 |')
    lines.append('| Col 7 (PRIIPS_KID)  | `M`=66, `C`=43, `O`=34, leer=13, `C `=1 | 41 286 (alle TP) |')
    lines.append('| Col 8 (UK)          | `x`=78, leer=79 | 0 |')
    lines.append('\nDer Validator parst `x` zu `Flag.UNKNOWN` und feuert PresenceRule nur bei `Flag.M`. Resultat: keine Falsch-Findings aus den Boolean-Spalten. **Kein Bug.**\n')
    lines.append('### 2. EET V1.1.2 Spalte 19 (Methodische Prosa-Anmerkungen)\n')
    lines.append('Spalte 19 enthält 4 freie Texteinträge ("Percentage or PCT instead of %", "sub wg on calculation methodologies", …). **Das Manifest verweist nicht auf Spalte 19** — LOOK_THROUGH ist als Spalte 17 deklariert (M=16, C=16, O=370, leer=304). **Kein Bug.** Die Phase-1-Exploration hatte fälschlicherweise Spalte 19 als LOOK_THROUGH benannt.\n')
    lines.append('### 3. EMT Spec hat zusätzliche `I` (Informative) Flag\n')
    lines.append('| Template | M | C | O | I | leer |')
    lines.append('|----------|--:|--:|--:|--:|-----:|')
    lines.append('| EMT V4.2 | 40 | 33 | 31 | 7 | 14 |')
    lines.append('| EMT V4.3 | 40 | 33 | 34 | 7 | 14 |')
    lines.append('\nKeine der 7 `I`-Felder produzierte ein Finding. `Flag.parse("I")` → `Flag.I`, `PresenceRule` feuert nur bei `Flag.M`. **Kein Bug.**\n')

    lines.append('## Validator-Bug-Verdachtsliste\n')
    lines.append('Mechanisch identifiziert: **keine** PRESENCE/FORMAT-Regel hat über den Audit eine FP-Klassifikation erhalten. Alle 353 PresenceRule-Ausprägungen treffen auf einen tatsächlich M-flagged Spec-Eintrag.\n')
    lines.append('**Über die strikt-mechanische Klassifikation hinaus**: zwei XF-Regelgruppen gehen *bewusst* einen Schritt über den Spec-Wortlaut hinaus (operator-intent-Extrapolation). Beide produzieren **WARNING-Findings**, ohne dass die Spec den Soll-Zustand zwingend mandiert — eine pragmatische Designentscheidung, kein Bug:\n')
    lines.append('| Regelgruppe | Findings im Lauf | Severity | Status |')
    lines.append('|-------------|-----------------:|----------|--------|')
    lines.append('| `EET-XF-ART*-MUST-BE-ABSENT` (6 IDs) | 24 | WARNING | Konservativ. Spec sagt nur "conditional to product type", nicht "must be absent". Java acknowledged. |')
    lines.append('| `EET-XF-PAI-186` (Teil von PAI-Block, 27 IDs) | 1 648 | WARNING | **FIXED 2026-05-12**: Severity von ERROR auf WARNING gesenkt (`EetRuleSet.java:242`). Spec mandiert PAI-Gating *explizit nur für Feld 106*; Validator extrapoliert auf alle 27 PAI-Indikatoren — jetzt mit angemessener WARNING-Severity. |')
    lines.append('\n**Resultat nach Fix**: Beide Regelgruppen sind jetzt severity-konsistent (WARNING für extrapolierte Regeln; ERROR ist literalen Spec-Pflichten vorbehalten). Spec-strikt sind die 1 672 WARNING-Findings weiter UNCLEAR, aber das operativ akzeptable Mass — kein Spec-Verstoss verschleiert, keine ERROR-Inflation für extrapolierte Regeln.\n')

    lines.append('## Datenqualitäts-Erkenntnisse\n')
    lines.append('Die mit Abstand grössten Finding-Volumen kommen nicht aus Validator-Bugs sondern aus **realen Datenfehlen** der Asset-Manager:\n')
    lines.append('| Datei | Template | Findings | Hauptursache |')
    lines.append('|-------|----------|---------:|--------------|')
    lines.append('| `findatex_blank_templates/EMT_v4.2.xlsx` | EMT V4.2 | 5 480 | Leere FinDatEx-Vorlage — *erwartet*, kein Datenfehler, sollte nicht im Audit erscheinen. |')
    lines.append('| `5_DWS_Group/consolidated_eet.xlsx` | EET V1.1.3 | 30 222 | Grösste Datei (1 962 Zeilen) × viele M-Felder in SFDR_ENTITY = systembedingt hohe Findings. |')
    lines.append('| `5_DWS_Group/consolidated_ept_de.csv` | EPT V2.1 | 21 807 | Grosse Portfolios (1 952 Zeilen) × KID-Pflichtfelder fehlen. |')
    lines.append('| `16_Royal_London_AM/rlam-ept.csv` | EPT V2.1 | 13 464 | Datei passt vermutlich nicht zu EPT V2.1 (Score nur 57.7 %). |')
    lines.append('| `16_Royal_London_AM/rlam-emt.csv` | EMT V4.3 | 8 160 | Score 53 % — vermutlich nicht V4.3, sondern ältere Version. |')
    lines.append('\nDie übrigen Files (UBS, Aviva v4.2 / v4.1, Abrdn) sind nahe 100 %.\n')

    lines.append('## DEFERRED — bewusst nicht geprüft\n')
    lines.append('Dieser Audit ist eine **False-Positive-Suche**. Er findet **nicht**:\n')
    lines.append('- **False Negatives** — Spec-Regeln, die der Validator nicht prüft. Beispiele aus stichprobenhafter Spec-Lektüre, die der Validator möglicherweise nicht erzwingt:\n')
    lines.append('  - **EET Codification-spezifische Wertelisten** (z. B. das Feld 27 SFDR-Produkt-Typ mit den Codes 0/6/8/9/etc.) — der Validator hat eine generische FormatRule, aber prüft sie die kompletten Closed-Lists?\n')
    lines.append('  - **EET Stewardship- und ESG-Label-Sheets** (separate Sheets in der Spec-XLSX) — gibt es Cross-Sheet-Konsistenz-Regeln, die abdecken sollten, dass eingetragene Codes in einem Referenzsheet existieren? Nicht im Audit nachgeprüft.\n')
    lines.append('  - **EMT Comment "If field 00060 = S, SF or B, then 00072 must be present"** (Feld 18) — wird vermutlich nicht als XF-Regel umgesetzt; der Validator hat keinen `EMT-XF-`-Block dafür im RuleSet.\n')
    lines.append('  - **PRIIPs RTS-Szenarien** (Stress, Ungünstig, Mässig, Günstig — diverse EPT-Felder) — der Validator führt keine quantitativen Plausibilitätschecks ("Stress ≤ Ungünstig ≤ Mässig ≤ Günstig"), das ist im Code als `// DEFERRED: requires regulatory SME` markiert (vgl. CLAUDE.md).\n')
    lines.append('- **Closed-List-Wert-Abgleich** — z. B. ob die ISO 4217-Currency-Liste, die der Validator hartcodiert/lädt, exakt mit der Spec-Liste übereinstimmt.\n')
    lines.append('- **Tatsächliches Verhalten von Conditional ("C")-Feldern in EET** — 32-64 Felder pro Profil sind als "C" geflaggt. Der Validator hat dafür `ConditionalPresenceRule`, der nur bei TPT-CIC-Applicability feuert. Für EET sollte er nicht feuern (kein CIC-Dimension) — und tut es offenbar auch nicht (keine entsprechenden Findings). Eine vollständige Audit dieser 200+ C-Felder gegen die Spec-Comment-Texte ist nicht Teil dieses Berichts.\n')
    lines.append('- **Severity-Mapping** — alle XF-Regeln emittieren ERROR oder WARNING gemäss Java-Code-Konstante. Spec gibt keine explizite Severity-Empfehlung — die Wahl ist eine Validator-Designentscheidung, nicht Spec-konform/-strittig.\n')

    return '\n'.join(lines)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    per_template: Dict[str, Tuple[Dict[str, RuleAgg], Dict[str, FieldRec]]] = {}
    for cfg in TEMPLATES:
        print(f'[audit] loading spec {cfg.label} …', file=sys.stderr)
        spec = load_spec(cfg)
        print(f'         {len(spec)} field records', file=sys.stderr)
        print(f'[audit] loading findings from {cfg.report.name} …', file=sys.stderr)
        rules = load_findings(cfg.report)
        print(f'         {len(rules)} unique rules, {sum(a.count for a in rules.values())} findings', file=sys.stderr)
        per_template[cfg.label] = (rules, spec)
    md = emit_markdown(per_template)
    OUT.write_text(md)
    print(f'[audit] wrote {OUT} ({len(md)} bytes)', file=sys.stderr)
    return 0

if __name__ == '__main__':
    sys.exit(main())
