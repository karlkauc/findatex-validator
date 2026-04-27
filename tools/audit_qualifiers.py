#!/usr/bin/env python3
"""Audit the CIC sub-category qualifier extraction against the real TPT V7 spec.

Reads every (qualifier_text, cic_class) cell from the spec, runs each through a
Python mirror of SpecLoader.parseSubcategoryQualifier, and compares against a
curated expectation map. For any (text, cic) not in the curated map, the
expected output is the empty set (no sub-category restriction).

Run:
  python3 tools/audit_qualifiers.py
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SPEC = ROOT / "src" / "main" / "resources" / "spec" / "TPT_V7  20241125_updated.xlsx"

CIC_COLS = list(range(12, 28))
CIC_NAMES = ["CIC0", "CIC1", "CIC2", "CIC3", "CIC4", "CIC5", "CIC6", "CIC7",
             "CIC8", "CIC9", "CICA", "CICB", "CICC", "CICD", "CICE", "CICF"]

QUOTED = re.compile(r'"([0-9A-Fa-f][0-9A-Za-z])"')
FOR_KW = re.compile(r"\bfor\b", re.IGNORECASE)
NON_ALNUM = re.compile(r"[^A-Za-z0-9]+")


def parse_subcategory_qualifier(text: str | None, cic_name: str) -> set[str]:
    """Mirror of the Java SpecLoader.parseSubcategoryQualifier."""
    if not text or not text.strip():
        return set()
    if not cic_name or not cic_name.startswith("CIC") or len(cic_name) != 4:
        return set()
    cls = cic_name[3].upper()

    subs: set[str] = set()
    for m in QUOTED.finditer(text):
        token = m.group(1).upper()
        if token[0] == cls:
            subs.add(token[1])

    forM = FOR_KW.search(text)
    if forM:
        tail = text[forM.end():]
        for tok in NON_ALNUM.split(tail):
            if len(tok) != 2:
                continue
            a = tok[0].upper()
            b = tok[1].upper()
            if a == cls and a.isalnum() and b.isalnum():
                subs.add(b)
    return subs


# Curated: qualifier text -> {cic class that produces a non-empty result -> expected sub-categories}.
# Any (text, cic) pair NOT listed here is expected to produce the empty set.
CURATED: dict[str, dict[str, set[str]]] = {
    # quoted patterns
    'x for convertible bonds "22" or other corporate bonds "29" quoted in units': {"CIC2": {"2", "9"}},
    'x for equity future "A1" and for commodity future "A5", other "A9"':         {"CICA": {"1", "5", "9"}},
    'x for equity options "B1", warrants "B4", commodities options "B5", others "B9"': {"CICB": {"1", "4", "5", "9"}},
    'x for equity options "C1", warrants "C4", commodities options "C5", others "C9"': {"CICC": {"1", "4", "5", "9"}},
    'x for equity legs of Total return swaps "D4", Security swaps "D5", others "D9"':  {"CICD": {"4", "5", "9"}},
    'x for interest rate future "A2", currency future "A3",  other "A9"':         {"CICA": {"2", "3", "9"}},
    'x for bond options "B2", currency options "B3", swaptions "B6", catastrohe and weather risk "B7", mortality risk "B8", other "B9"': {"CICB": {"2", "3", "6", "7", "8", "9"}},
    'x for bond options "C2", currency options "C3", swaptions "C6", catastrohe and weather risk "C7", mortality risk "C8", other "C9"': {"CICC": {"2", "3", "6", "7", "8", "9"}},
    # unquoted patterns
    "x for 22":                     {"CIC2": {"2"}},
    "x for D4, D5":                 {"CICD": {"4", "5"}},
    "x for D1, D3":                 {"CICD": {"1", "3"}},
    "x for F1, F3, F4":             {"CICF": {"1", "3", "4"}},
    "x for F1, F2":                 {"CICF": {"1", "2"}},
    "x for E1":                     {"CICE": {"1"}},
    "x for A1":                     {"CICA": {"1"}},
    "x for B1, B4":                 {"CICB": {"1", "4"}},
    "x for C1, C4,":                {"CICC": {"1", "4"}},
    "x for C1, C4, B1, B4":         {"CICB": {"1", "4"}, "CICC": {"1", "4"}},
    "x for B1, B4, C1, C4":         {"CICB": {"1", "4"}, "CICC": {"1", "4"}},
    "x for A2":                     {"CICA": {"2"}},
    "x for B2, C2":                 {"CICB": {"2"}, "CICC": {"2"}},
    "x for B2":                     {"CICB": {"2"}},
    "x for C2":                     {"CICC": {"2"}},
    "x for A3":                     {"CICA": {"3"}},
    "x for B3":                     {"CICB": {"3"}},
    "x for C3":                     {"CICC": {"3"}},
    "x\nfor E2":                    {"CICE": {"2"}},
    "x for A3, A5":                 {"CICA": {"3", "5"}},
    "x for 73, 74, 75":             {"CIC7": {"3", "4", "5"}},
    "x for 73,74,75":               {"CIC7": {"3", "4", "5"}},
    "x for 52, 54":                 {"CIC5": {"2", "4"}},
    "x for 62,64":                  {"CIC6": {"2", "4"}},
    "x for 51, 53, 56":             {"CIC5": {"1", "3", "6"}},
    "x for 61,63, 66":              {"CIC6": {"1", "3", "6"}},
}

# Patterns that we explicitly recognise as cross-field conditional text — not
# sub-category restrictions. They should always produce the empty set under any CIC.
KNOWN_CROSSFIELD_TEXTS = {
    'x\nif item 34 is not blank',
    'x\nif item 116 set to "1"',
    'x\nif item 120 set to "1"',
    'x\nif item 48 set to "1"',
    'x\nif item 51 set to "1"',
    'x\nif item 32 set to "Floating"',
    'x\nif item 42 is not blank',
    'x\nif item 85 set to "1"',
    'x if 138 is "1" "2" or "3"',
    'x only if 134 is set to 1',
    'x\nIf item 29 is not blank',
    'if item 42 is Equal to Cal or Put',
    'If coming from the lookthrough of an underlying fund',
    'x for all legs of all swaps',  # prose, no sub-codes
    'M',
}


def expected_for(text: str, cic: str) -> set[str]:
    return CURATED.get(text, {}).get(cic, set())


def collect_spec_pairs() -> list[tuple[str, str, int, str]]:
    """Return [(qualifier_text, cic_class, source_row, num_data), ...]."""
    wb = openpyxl.load_workbook(SPEC, data_only=True)
    ws = wb["TPT V7.0"]
    out = []
    for r in range(8, ws.max_row + 1):
        num = ws.cell(row=r, column=1).value
        path = ws.cell(row=r, column=2).value
        if not num:
            continue
        # Skip section-header rows (no numeric prefix in NUM_DATA, no path)
        snum = str(num)
        if not path:
            head = snum.split("_", 1)[0]
            if not any(c.isdigit() for c in head):
                continue
        for ci, col in enumerate(CIC_COLS):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            s = str(v).strip()
            if not s or s.lower() == "x":
                continue
            out.append((s, CIC_NAMES[ci], r, snum))
    return out


def main() -> int:
    pairs = collect_spec_pairs()
    print(f"Spec file: {SPEC}")
    print(f"Non-trivial qualifier cells in spec: {len(pairs)}\n")

    failures: list[tuple[str, str, set[str], set[str], int, str]] = []
    unknown_texts: dict[str, list[tuple[str, int, str]]] = {}

    for text, cic, row, num in pairs:
        actual = parse_subcategory_qualifier(text, cic)
        expected = expected_for(text, cic)
        if actual != expected:
            failures.append((text, cic, expected, actual, row, num))
        # Track texts we have no curated entry for and that aren't in the
        # known-crossfield list — these need explicit decisions.
        if text not in CURATED and text not in KNOWN_CROSSFIELD_TEXTS:
            unknown_texts.setdefault(text, []).append((cic, row, num))

    if failures:
        print(f"  ✗ {len(failures)} mismatches:")
        for text, cic, exp, act, row, num in failures:
            print(f"      row {row:>3}  {cic}  {num[:40]:<42s}  exp={sorted(exp) or '∅'!s}  got={sorted(act) or '∅'!s}  text={text!r}")
    else:
        print(f"  ✓ All {len(pairs)} spec qualifier cells produce the curated expected output.")

    if unknown_texts:
        print(f"\n  ⚠ {len(unknown_texts)} qualifier text(s) outside both the curated table and the known-crossfield set:")
        for text, occ in sorted(unknown_texts.items()):
            print(f"      {text!r}  (used in {len(occ)} cell{'s' if len(occ) > 1 else ''})")
    else:
        print(f"\n  ✓ Every distinct qualifier text in the spec is classified.")

    return 0 if not failures and not unknown_texts else 1


if __name__ == "__main__":
    sys.exit(main())
