#!/usr/bin/env python3
"""Generate 10 example TPT V7 files demonstrating different validation outcomes.

  01_clean.xlsx                       — passes every rule (overall score ~ 100 %).
  02_missing_mandatory.csv            — drops several M-flagged fields.
  03_bad_formats.xlsx                 — invalid currency, country, ISO date, NACE.
  04_bad_closed_lists.xlsx            — values outside the closed-list codifications.
  05_bad_isin_checksum.xlsx           — instrument code with wrong Luhn check digit.
  06_bad_lei_checksum.xlsx            — issuer code with wrong ISO 17442 check.
  07_weights_dont_sum.xlsx            — Σ field 26 PositionWeight far from 1.
  08_nav_mismatch.xlsx                — TotalNetAssets disagrees with SharePrice × Shares
                                        and CashPercentage disagrees with the cash CIC sum.
  09_interest_rate_inconsistent.xlsx  — Floating bond missing index/margin; fixed bond
                                        missing coupon rate.
  10_dates_and_derivatives.xlsx       — reporting < valuation, maturity in past,
                                        futures without underlying CIC, PIK code on equity.

Outputs land in samples/tpt/ at the project root.
"""
from __future__ import annotations

import csv
import json
from copy import deepcopy
from pathlib import Path

import openpyxl

import tpt_showcase

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "samples" / "tpt"
OUT.mkdir(parents=True, exist_ok=True)

SPEC_MANIFEST = ROOT / "core/src/main/resources/spec/tpt/tpt-v7-info.json"
SPEC_XLSX = ROOT / "core/src/main/resources/spec/tpt/TPT_V7_20241125.xlsx"

HEADERS = [
    "1_Portfolio_identifying_data",
    "2_Type_of_identification_code_for_the_fund_share_or_portfolio",
    "3_Portfolio_name",
    "4_Portfolio_currency_(B)",
    "5_Net_asset_valuation_of_the_portfolio_or_the_share_class_in_portfolio_currency",
    "6_Valuation_date",
    "7_Reporting_date",
    "8_Share_price",
    "8b_Total_number_of_shares",
    "9_Cash_ratio",
    "11_Complete_SCR_delivery",
    "12_CIC_code_of_the_instrument",
    "14_Identification_code_of_the_instrument",
    "15_Type_of_identification_code_for_the_instrument",
    "17_Instrument_name",
    "21_Quotation_currency_(A)",
    "22_Market_valuation_in_quotation_currency_(A)",
    "23_Clean_market_valuation_in_quotation_currency_(A)",
    "24_Market_valuation_in_portfolio_currency_(B)",
    "25_Clean_market_valuation_in_portfolio_currency_(B)",
    "26_Valuation_weight",
    "32_Interest_rate_type",
    "33_Coupon_rate",
    "34_Interest_rate_reference_identification",
    "35_Identification_type_for_interest_rate_index",
    "36_Interest_rate_index_name",
    "37_Interest_rate_margin",
    "38_Coupon_payment_frequency",
    "39_Maturity_date",
    "40_Redemption_type",
    "41_Redemption_rate",
    "46_Issuer_name",
    "47_Issuer_identification_code",
    "48_Type_of_identification_code_for_issuer",
    "52_Issuer_country",
    "64_Exercise_type",
    "67_CIC_of_the_underlying_asset",
    "131_Underlying_asset_category",
    "146_PIK",
    "1000_TPT_Version",
]


def spec_headers(nums: list[str]) -> list[str]:
    """Turn field numbers into the exact header strings the spec itself uses.

    HeaderMapper matches on the token before the first underscore
    (SpecCatalog.matchHeader), so only the number has to be right — but taking
    the whole string from the spec sheet keeps the showcase readable and
    guarantees it can never drift from the field it claims to be.
    """
    manifest = json.loads(SPEC_MANIFEST.read_text(encoding="utf-8"))
    cols = manifest["columns"]
    sheet = openpyxl.load_workbook(SPEC_XLSX, data_only=True)[manifest["sheetName"]]
    # In the TPT manifests "numData" is the column holding the "<num>_<label>"
    # token (the header form); "path" is the dotted XML path. The EET/EMT/EPT
    # manifests split those differently — see findatex_sample_helpers.load_spec.
    by_num: dict[str, str] = {}
    for r in range(manifest["firstDataRow"], sheet.max_row + 1):
        token = sheet.cell(r, cols["numData"]).value
        if token is None:
            continue
        token = str(token).strip()
        num, sep, _ = token.partition("_")
        # Section headers ("Position information") have no "<num>_" prefix;
        # this mirrors SpecLoader.looksLikeFieldLabel.
        if not sep or not any(ch.isdigit() for ch in num):
            continue
        by_num.setdefault(num, token)
    missing = [n for n in nums if n not in by_num]
    if missing:
        raise SystemExit(f"build_examples: no TPT V7 field for {missing}")
    return [by_num[n] for n in nums]


def base_row(numdata: dict) -> dict:
    """Fill all headers with empty strings then overlay the supplied keys."""
    row = {h: "" for h in HEADERS}
    row.update(numdata)
    return row


PORT = {
    "1_Portfolio_identifying_data": "FR0010000001",
    "2_Type_of_identification_code_for_the_fund_share_or_portfolio": "1",
    "3_Portfolio_name": "Demo Bond Fund",
    "4_Portfolio_currency_(B)": "EUR",
    "5_Net_asset_valuation_of_the_portfolio_or_the_share_class_in_portfolio_currency": "10000000",
    "6_Valuation_date": "2025-12-31",
    "7_Reporting_date": "2025-12-31",
    "8_Share_price": "100",
    "8b_Total_number_of_shares": "100000",
    "9_Cash_ratio": "0.20",
    "11_Complete_SCR_delivery": "N",
    "1000_TPT_Version": "V7.0 (official) dated 25 November 2024",
}

GOV_BOND = base_row({
    **PORT,
    "12_CIC_code_of_the_instrument": "FR12",
    "14_Identification_code_of_the_instrument": "FR0000571085",  # valid ISIN
    "15_Type_of_identification_code_for_the_instrument": "1",
    "17_Instrument_name": "FR Treasury 1.5% 2030",
    "21_Quotation_currency_(A)": "EUR",
    "22_Market_valuation_in_quotation_currency_(A)": "5000000",
    "23_Clean_market_valuation_in_quotation_currency_(A)": "4900000",
    "24_Market_valuation_in_portfolio_currency_(B)": "5000000",
    "25_Clean_market_valuation_in_portfolio_currency_(B)": "4900000",
    "26_Valuation_weight": "0.5",
    "32_Interest_rate_type": "Fixed",
    "33_Coupon_rate": "0.015",
    "38_Coupon_payment_frequency": "1",
    "39_Maturity_date": "2030-05-25",
    "40_Redemption_type": "Bullet",
    "41_Redemption_rate": "1",
    "46_Issuer_name": "French Republic",
    "47_Issuer_identification_code": "969500TJ5KRTCJQSU990",  # synthetic but checksum-valid LEI (French Treasury-flavoured)
    "48_Type_of_identification_code_for_issuer": "1",
    "52_Issuer_country": "FR",
    "131_Underlying_asset_category": "1",
})

EQUITY = base_row({
    **PORT,
    "12_CIC_code_of_the_instrument": "DE31",
    "14_Identification_code_of_the_instrument": "DE0007164600",  # valid ISIN (SAP)
    "15_Type_of_identification_code_for_the_instrument": "1",
    "17_Instrument_name": "SAP SE",
    "21_Quotation_currency_(A)": "EUR",
    "22_Market_valuation_in_quotation_currency_(A)": "3000000",
    "23_Clean_market_valuation_in_quotation_currency_(A)": "3000000",
    "24_Market_valuation_in_portfolio_currency_(B)": "3000000",
    "25_Clean_market_valuation_in_portfolio_currency_(B)": "3000000",
    "26_Valuation_weight": "0.3",
    "46_Issuer_name": "SAP SE",
    "47_Issuer_identification_code": "529900D6BF99LW9R2E68",  # valid LEI (SAP)
    "48_Type_of_identification_code_for_issuer": "1",
    "52_Issuer_country": "DE",
    "131_Underlying_asset_category": "3L",
})

CASH = base_row({
    **PORT,
    "12_CIC_code_of_the_instrument": "XL71",
    "14_Identification_code_of_the_instrument": "CASH-EUR-001",
    "15_Type_of_identification_code_for_the_instrument": "99",
    "17_Instrument_name": "Cash account EUR",
    "21_Quotation_currency_(A)": "EUR",
    "22_Market_valuation_in_quotation_currency_(A)": "2000000",
    "23_Clean_market_valuation_in_quotation_currency_(A)": "2000000",
    "24_Market_valuation_in_portfolio_currency_(B)": "2000000",
    "25_Clean_market_valuation_in_portfolio_currency_(B)": "2000000",
    "26_Valuation_weight": "0.2",
    "46_Issuer_name": "Demo Custodian Bank",
    "131_Underlying_asset_category": "7",
})

CLEAN_ROWS = [GOV_BOND, EQUITY, CASH]


PORT_B = {
    **PORT,
    "1_Portfolio_identifying_data": "DE0010000002",
    "3_Portfolio_name": "Demo Equity Fund DE",
}

PORT_C = {
    **PORT,
    "1_Portfolio_identifying_data": "LU0010000003",
    "3_Portfolio_name": "Demo Mixed Fund LU",
}


def _swap_port(rows: list[dict], port: dict) -> list[dict]:
    """Return a deep-copied row list with all PORT-level fields swapped to `port`."""
    return [{**r, **port} for r in rows]


# --------- helpers to build identifiers that pass the local checksum ------

def _alpha_to_digits(s: str) -> str:
    out = []
    for ch in s:
        if ch.isdigit():
            out.append(ch)
        else:
            out.append(str(ord(ch.upper()) - ord("A") + 10))
    return "".join(out)


def isin_check_digit(body11: str) -> str:
    """Luhn check digit for an 11-char ISIN body (2 country letters + 9 alnum)."""
    digits = _alpha_to_digits(body11)
    total, double_it = 0, True
    for d in reversed(digits):
        n = int(d)
        if double_it:
            n *= 2
            if n > 9:
                n -= 9
        total += n
        double_it = not double_it
    return str((10 - (total % 10)) % 10)


def lei_check_digits(body18: str) -> str:
    """ISO 17442 mod-97 check digits for an 18-char LEI body."""
    n = int(_alpha_to_digits(body18 + "00")) % 97
    return f"{98 - n:02d}"


def make_isin(body11: str) -> str:
    return body11 + isin_check_digit(body11)


def make_lei(body18: str) -> str:
    return body18 + lei_check_digits(body18)


# ---------------------------------------------------------------- writers --

def write_xlsx(path: Path, rows: list[dict], headers: list[str] | None = None) -> None:
    headers = headers or HEADERS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TPT V7"
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h
    for r, row in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            ws.cell(row=r, column=c).value = row.get(h, "")
    wb.save(path)
    print(f"Wrote {path.relative_to(ROOT)}")


def write_csv(path: Path, rows: list[dict], delimiter: str = ";") -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, delimiter=delimiter)
        w.writerow(HEADERS)
        for row in rows:
            w.writerow([row.get(h, "") for h in HEADERS])
    print(f"Wrote {path.relative_to(ROOT)}")


# --------------------------------------------------------------- scenarios -

def s00_showcase() -> list[dict]:
    """The file behind "Try an example": 3 funds, 60 positions, ~90 fields.

    Rows come back keyed by field number; ``main()`` projects them onto the
    spec header strings. Identifiers are verified here rather than in
    tpt_showcase so the module stays free of imports.
    """
    tpt_showcase.assert_identifiers(isin_check_digit, lei_check_digits)
    nums = tpt_showcase.SHOWCASE_NUMS
    headers = spec_headers(nums)
    return [{h: row[n] for n, h in zip(nums, headers)}
            for row in tpt_showcase.build_showcase()]


def s01_clean() -> list[dict]:
    return deepcopy(CLEAN_ROWS)


def s02_missing_mandatory() -> list[dict]:
    rows = deepcopy(CLEAN_ROWS)
    # drop M fields on the first row
    rows[0]["5_Net_asset_valuation_of_the_portfolio_or_the_share_class_in_portfolio_currency"] = ""
    rows[0]["6_Valuation_date"] = ""
    rows[0]["12_CIC_code_of_the_instrument"] = ""
    rows[1]["14_Identification_code_of_the_instrument"] = ""
    rows[1]["17_Instrument_name"] = ""
    rows[2]["4_Portfolio_currency_(B)"] = ""
    return rows


def s03_bad_formats() -> list[dict]:
    rows = deepcopy(CLEAN_ROWS)
    rows[0]["21_Quotation_currency_(A)"] = "ZZZ"     # invalid ISO 4217
    rows[0]["52_Issuer_country"] = "XX"               # invalid ISO 3166
    rows[0]["6_Valuation_date"] = "31/12/2025"        # wrong date format
    rows[0]["7_Reporting_date"] = "2025-13-40"        # impossible date
    rows[1]["52_Issuer_country"] = "GERMANY"          # not 2-letter
    rows[2]["4_Portfolio_currency_(B)"] = "Eu"        # too short
    rows[2]["8_Share_price"] = "abc"                  # not numeric
    return rows


def s04_bad_closed_lists() -> list[dict]:
    rows = deepcopy(CLEAN_ROWS)
    # Field 15 closed list = {1..9, 99}
    rows[0]["15_Type_of_identification_code_for_the_instrument"] = "42"
    # Field 38 (coupon frequency) ∈ {0,1,2,4,12,52}
    rows[0]["38_Coupon_payment_frequency"] = "3"
    # Field 40 (redemption type) ∈ {Bullet, Sinkable, ...}
    rows[0]["40_Redemption_type"] = "VeryWeirdRedemption"
    # Field 11 only allows Y/N
    rows[0]["11_Complete_SCR_delivery"] = "Maybe"
    # Field 64 (exercise type) ∈ {AM, EU, AS, BE} — set on the equity row even
    # though it usually applies to options, just to trigger the closed-list rule.
    rows[1]["64_Exercise_type"] = "QQ"
    # Field 131 (underlying asset category) closed list — give it a bogus code
    rows[2]["131_Underlying_asset_category"] = "Z9"
    return rows


def s05_bad_isin_checksum() -> list[dict]:
    rows = deepcopy(CLEAN_ROWS)
    # Flip the last digit so the Luhn check fails
    rows[0]["14_Identification_code_of_the_instrument"] = "FR0000571086"  # was 5, now 6
    rows[1]["14_Identification_code_of_the_instrument"] = "DE0007164601"  # was 0, now 1
    return rows


def s06_bad_lei_checksum() -> list[dict]:
    rows = deepcopy(CLEAN_ROWS)
    # Flip the last digit so mod-97 != 1
    rows[0]["47_Issuer_identification_code"] = "969500TJ5KRTCJQSU991"  # off-by-one in check digits
    rows[1]["47_Issuer_identification_code"] = "529900D6BF99LW9R2E69"
    return rows


def s07_weights_dont_sum() -> list[dict]:
    rows = deepcopy(CLEAN_ROWS)
    # weights add up to 0.7 instead of 1.0 (well outside the ±2 % tolerance)
    rows[0]["26_Valuation_weight"] = "0.3"
    rows[1]["26_Valuation_weight"] = "0.2"
    rows[2]["26_Valuation_weight"] = "0.2"
    return rows


def s08_nav_mismatch() -> list[dict]:
    rows = deepcopy(CLEAN_ROWS)
    # SharePrice × Shares = 100 × 100 000 = 10 000 000 = TotalNetAssets in clean.
    # Bump the share price so the product is far off.
    for r in rows:
        r["8_Share_price"] = "150"             # 150 × 100 000 = 15 000 000 ≠ 10 000 000
        r["9_Cash_ratio"] = "0.50"             # declared cash 50 %; actual cash share = 20 %
    return rows


def s09_interest_rate_inconsistent() -> list[dict]:
    rows = deepcopy(CLEAN_ROWS)
    # Govt bond: mark Floating but leave 34..37 empty -> XF-10 fires for each missing field.
    rows[0]["32_Interest_rate_type"] = "Floating"
    rows[0]["33_Coupon_rate"] = ""
    rows[0]["34_Interest_rate_reference_identification"] = ""
    rows[0]["35_Identification_type_for_interest_rate_index"] = ""
    rows[0]["36_Interest_rate_index_name"] = ""
    rows[0]["37_Interest_rate_margin"] = ""
    # Add a second bond marked Fixed but with no coupon rate.
    extra = base_row({
        **PORT,
        "12_CIC_code_of_the_instrument": "DE22",
        "14_Identification_code_of_the_instrument": "DE000A30VYR2",  # placeholder code
        "15_Type_of_identification_code_for_the_instrument": "99",
        "17_Instrument_name": "Mystery Corp 2% 2029",
        "21_Quotation_currency_(A)": "EUR",
        "22_Market_valuation_in_quotation_currency_(A)": "1000000",
        "23_Clean_market_valuation_in_quotation_currency_(A)": "990000",
        "24_Market_valuation_in_portfolio_currency_(B)": "1000000",
        "25_Clean_market_valuation_in_portfolio_currency_(B)": "990000",
        "26_Valuation_weight": "0.1",
        "32_Interest_rate_type": "Fixed",
        "33_Coupon_rate": "",                  # ← XF-10 trigger
        "38_Coupon_payment_frequency": "2",
        "39_Maturity_date": "2029-09-30",
        "40_Redemption_type": "Bullet",
        "41_Redemption_rate": "1",
        "46_Issuer_name": "Mystery Corp",
        "47_Issuer_identification_code": "",
        "48_Type_of_identification_code_for_issuer": "9",
        "52_Issuer_country": "DE",
        "131_Underlying_asset_category": "2",
    })
    rows.append(extra)
    # Fix weight sum so XF-04 stays clean.
    rows[0]["26_Valuation_weight"] = "0.4"
    rows[1]["26_Valuation_weight"] = "0.3"
    rows[2]["26_Valuation_weight"] = "0.2"
    extra["26_Valuation_weight"] = "0.1"
    return rows


def s11_unknown_isin_lei() -> list[dict]:
    """Synthetic but checksum-valid ISINs/LEIs that don't exist in the
    public registers — local validation passes, online lookup (GLEIF /
    OpenFIGI) flags ``LEI-LIVE/...`` and ``ISIN-LIVE/...`` ERRORs.

    Also exercises the optional WARNING sub-toggles when enabled in
    Settings: row 1 carries an issuer name that mismatches the GLEIF
    record (none — the LEI is unknown), and a quotation currency that
    differs from any record OpenFIGI would return.
    """
    rows = deepcopy(CLEAN_ROWS)

    # Row 0 (gov bond): replace ISIN + LEI with synthetic but checksum-valid codes.
    # XS = Eurobond pseudo-country, valid ISIN format. Body 999999999 is unallocated.
    fake_isin_a = make_isin("XS999999999")
    fake_lei_a = make_lei("ZZ0000000000ZZZZZZ")
    rows[0]["14_Identification_code_of_the_instrument"] = fake_isin_a
    rows[0]["47_Issuer_identification_code"] = fake_lei_a
    rows[0]["46_Issuer_name"] = "Synthetic Issuer Plc"

    # Row 1 (equity): different synthetic pair so the file shows two distinct
    # online ERRORs (one per service per identifier kind).
    fake_isin_b = make_isin("XS888888888")
    fake_lei_b = make_lei("YY1111111111YYYYYY")
    rows[1]["14_Identification_code_of_the_instrument"] = fake_isin_b
    rows[1]["47_Issuer_identification_code"] = fake_lei_b
    rows[1]["46_Issuer_name"] = "Imaginary Holdings AG"

    # Row 2 (cash) keeps "CASH-EUR-001" with type 99 — not an ISIN, untouched.
    return rows


def s10_dates_and_derivatives() -> list[dict]:
    rows = deepcopy(CLEAN_ROWS)
    # Reporting < valuation -> XF-12
    for r in rows:
        r["6_Valuation_date"] = "2025-12-31"
        r["7_Reporting_date"] = "2025-11-30"
    # Maturity in the past for the gov bond -> XF-11
    rows[0]["39_Maturity_date"] = "2020-01-01"
    # Add a futures position (CIC A) without underlying CIC -> XF-14
    futures = base_row({
        **PORT,
        "6_Valuation_date": "2025-12-31",
        "7_Reporting_date": "2025-11-30",
        "12_CIC_code_of_the_instrument": "XLA1",
        "14_Identification_code_of_the_instrument": "FUT-EUREX-001",
        "15_Type_of_identification_code_for_the_instrument": "99",
        "17_Instrument_name": "EUREX FBund Future Mar26",
        "21_Quotation_currency_(A)": "EUR",
        "22_Market_valuation_in_quotation_currency_(A)": "0",
        "23_Clean_market_valuation_in_quotation_currency_(A)": "0",
        "24_Market_valuation_in_portfolio_currency_(B)": "0",
        "25_Clean_market_valuation_in_portfolio_currency_(B)": "0",
        "26_Valuation_weight": "0.0",
        "67_CIC_of_the_underlying_asset": "",       # ← XF-14 trigger
        "131_Underlying_asset_category": "A",
    })
    rows.append(futures)
    # Put a PIK code on the equity row (PIK only meaningful for bonds/loans) -> XF-13 warning
    rows[1]["146_PIK"] = "2"
    return rows


def s12_multi_fund_clean() -> list[dict]:
    """Three distinct funds (FR, DE, LU), each with a clean 3-position book."""
    return (deepcopy(CLEAN_ROWS)
            + _swap_port(deepcopy(CLEAN_ROWS), PORT_B)
            + _swap_port(deepcopy(CLEAN_ROWS), PORT_C))


def s13_multi_fund_with_errors() -> list[dict]:
    """Three funds with errors distributed across them — Fund A clean, Fund B
    has weight-sum and ISIN-checksum bugs, Fund C is missing NAV."""
    a = deepcopy(CLEAN_ROWS)
    b = _swap_port(deepcopy(CLEAN_ROWS), PORT_B)
    # Fund B: weights 0.3 + 0.2 + 0.2 = 0.7 (XF-04 fires for B only)
    b[0]["26_Valuation_weight"] = "0.3"
    b[1]["26_Valuation_weight"] = "0.2"
    b[2]["26_Valuation_weight"] = "0.2"
    # Fund B: corrupt ISIN checksum on first position (ISIN/14 fires for B only)
    b[0]["14_Identification_code_of_the_instrument"] = "FR0000571086"
    c = _swap_port(deepcopy(CLEAN_ROWS), PORT_C)
    # Fund C: drop the mandatory NAV (field 5) on the first row
    c[0]["5_Net_asset_valuation_of_the_portfolio_or_the_share_class_in_portfolio_currency"] = ""
    return a + b + c


# ----------------------------------------------------------------- driver --

EXAMPLES = [
    ("00_showcase.xlsx",                    s00_showcase,               "showcase"),
    ("01_clean.xlsx",                       s01_clean,                  "xlsx"),
    ("02_missing_mandatory.csv",            s02_missing_mandatory,      "csv"),
    ("03_bad_formats.xlsx",                 s03_bad_formats,            "xlsx"),
    ("04_bad_closed_lists.xlsx",            s04_bad_closed_lists,       "xlsx"),
    ("05_bad_isin_checksum.xlsx",           s05_bad_isin_checksum,      "xlsx"),
    ("06_bad_lei_checksum.xlsx",            s06_bad_lei_checksum,       "xlsx"),
    ("07_weights_dont_sum.xlsx",            s07_weights_dont_sum,       "xlsx"),
    ("08_nav_mismatch.xlsx",                s08_nav_mismatch,           "xlsx"),
    ("09_interest_rate_inconsistent.xlsx",  s09_interest_rate_inconsistent, "xlsx"),
    ("10_dates_and_derivatives.xlsx",       s10_dates_and_derivatives,  "xlsx"),
    ("11_unknown_isin_lei.xlsx",            s11_unknown_isin_lei,       "xlsx"),
    ("12_multi_fund_clean.xlsx",            s12_multi_fund_clean,       "xlsx"),
    ("13_multi_fund_with_errors.xlsx",      s13_multi_fund_with_errors, "xlsx"),
]


def main() -> int:
    for filename, factory, fmt in EXAMPLES:
        path = OUT / filename
        rows = factory()
        if fmt == "csv":
            write_csv(path, rows)
        elif fmt == "showcase":
            write_xlsx(path, rows, spec_headers(tpt_showcase.SHOWCASE_NUMS))
        else:
            write_xlsx(path, rows)

    # Generate a small README inside samples/tpt/ documenting each file.
    readme = OUT / "README.md"
    lines = [
        "# Example TPT V7 files",
        "",
        "Auto-generated by `tools/build_examples.py`. The showcase file is the",
        "demo behind *\"No file at hand? Try an example\"* on the web app; the",
        "numbered files are one-rule-family-each regression fixtures.",
        "",
        "| File | What it demonstrates |",
        "|------|----------------------|",
        "| `00_showcase.xlsx` | **The demo file.** Three funds \u00d7 20 positions (60 rows, ~90 spec fields): government and corporate bonds, equities, ETFs, a Bund future, an index option and cash. NAV, position weights, share price and cash ratio are computed from the book, so they reconcile. Fund A is clean; fund B carries data-entry errors (ISIN checksum, LEI checksum, `EURO` as a currency, `GERMANY` as a country, an out-of-list code type, missing instrument code and name); fund C carries cross-field inconsistencies (XF-04 weight sum, XF-05 cash ratio, XF-08 coupon frequency, XF-09 custodian pair, XF-10 floater without index, XF-11 past maturity, XF-14 future without underlying CIC). |",
        "| `01_clean.xlsx` | Fully valid file (overall score ≈ 100 %). |",
        "| `02_missing_mandatory.csv` | Missing M-flagged fields (5, 6, 12, 14, 17, 4) → PRESENCE errors. |",
        "| `03_bad_formats.xlsx` | Invalid ISO 4217 currency, ISO 3166 country, ISO 8601 date, NACE code, numeric. |",
        "| `04_bad_closed_lists.xlsx` | Values outside closed lists for fields 11, 15, 38, 40, 64, 131. |",
        "| `05_bad_isin_checksum.xlsx` | Instrument codes with corrupt Luhn check digit. |",
        "| `06_bad_lei_checksum.xlsx` | Issuer LEIs with corrupt ISO 17442 mod-97 check. |",
        "| `07_weights_dont_sum.xlsx` | Σ field 26 = 0.7 (outside ±2 % tolerance) → XF-04. |",
        "| `08_nav_mismatch.xlsx` | TotalNetAssets ≠ SharePrice × Shares; CashPercentage off → XF-05, XF-06. |",
        "| `09_interest_rate_inconsistent.xlsx` | Floating bond missing index/margin; Fixed bond missing coupon rate → XF-10. |",
        "| `10_dates_and_derivatives.xlsx` | Reporting < Valuation; maturity in past; futures without underlying CIC; PIK on equity → XF-11, XF-12, XF-13, XF-14. |",
        "| `11_unknown_isin_lei.xlsx` | Synthetic ISINs/LEIs that pass the local Luhn/mod-97 check but aren't registered in OpenFIGI/GLEIF. Local validation runs clean; **enable Online validation in Settings** to see `LEI-LIVE/...` and `ISIN-LIVE/...` ERRORs. |",
        "| `12_multi_fund_clean.xlsx` | Three distinct funds (FR / DE / LU) in one file, each with a clean 3-position book → no errors expected. |",
        "| `13_multi_fund_with_errors.xlsx` | Three funds with errors distributed: Fund B has a weight-sum mismatch (XF-04) and an ISIN checksum bug (ISIN/14); Fund C is missing the mandatory NAV (PRESENCE/5). Each finding must carry the correct portfolio context. |",
        "",
        "",
        "Equity, ETF and OAT ISINs are the real, publicly published ones. Bond and",
        "derivative ISINs are constructed: checksum-valid and in a plausible",
        "national range, but not registered. LEIs are constructed except SAP's, so",
        "they pass the mod-97 check but will not resolve against GLEIF — expected,",
        "not a bug, when online validation is switched on.",
        "",
        "Open them via the JavaFX UI (`mvn javafx:run` → *Browse…*) to see the",
        "rule engine catch each issue, or feed them through the JUnit suite.",
        "",
    ]
    readme.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {readme.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
