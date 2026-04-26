#!/usr/bin/env python3
"""Generate three synthetic TPT V7 sample files used by the JUnit suite.

  - clean_v7.xlsx        — minimal but valid: one government bond, one equity, one cash.
  - missing_mandatory.csv — same data but with field 5 (TotalNetAssets) and field 6 (ValuationDate) removed.
  - bad_formats.xlsx      — identical to clean but with bogus currency & coupon-frequency values.

The sample files are written to src/test/resources/sample/.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "src" / "test" / "resources" / "sample"
OUT.mkdir(parents=True, exist_ok=True)

# Pick a curated subset of mandatory fields that cover the validation paths we care about.
HEADERS = [
    "1_Portfolio_identifying_data",
    "2_Type_of_identification_code_for_the_fund_share_or_portfolio",
    "3_Portfolio_name",
    "4_Portfolio_currency_(B)",
    "5_Net_asset_valuation_of_the_portfolio_or_the_share_class_in_portfolio_currency",
    "6_Valuation_date",
    "7_Reporting_date",
    "8_Share_price",
    "8b_Total_number_of_shares",
    "11_Complete_SCR_delivery",
    "12_CIC_code_of_the_instrument",
    "14_Identification_code_of_the_instrument",
    "15_Type_of_identification_code_for_the_instrument",
    "17_Instrument_name",
    "21_Quotation_currency_(A)",
    "22_Market_valuation_in_quotation_currency_(A)",
    "23_Clean_market_valuation_in_quotation_currency_(A)",
    "24_Market_valuation_in_portfolio_currency_(B)",
    "25_Clean_market_valuation_in_portfolio_currency_(B)",
    "26_Valuation_weight",
    "32_Interest_rate_type",
    "33_Coupon_rate",
    "38_Coupon_payment_frequency",
    "39_Maturity_date",
    "40_Redemption_type",
    "41_Redemption_rate",
    "46_Issuer_name",
    "47_Issuer_identification_code",
    "48_Type_of_identification_code_for_issuer",
    "131_Underlying_asset_category",
    "1000_TPT_Version",
]

ROWS_CLEAN = [
    # CIC 1: Government bond
    {
        "1_Portfolio_identifying_data": "FR0010000001",
        "2_Type_of_identification_code_for_the_fund_share_or_portfolio": "1",
        "3_Portfolio_name": "Demo Bond Fund",
        "4_Portfolio_currency_(B)": "EUR",
        "5_Net_asset_valuation_of_the_portfolio_or_the_share_class_in_portfolio_currency": "10000000",
        "6_Valuation_date": "2025-12-31",
        "7_Reporting_date": "2025-12-31",
        "8_Share_price": "100",
        "8b_Total_number_of_shares": "100000",
        "11_Complete_SCR_delivery": "N",
        "12_CIC_code_of_the_instrument": "FR12",
        "14_Identification_code_of_the_instrument": "FR0000571085",  # valid ISIN (FR Treasury, real)
        "15_Type_of_identification_code_for_the_instrument": "1",
        "17_Instrument_name": "FR Treasury 1.5% 2030",
        "21_Quotation_currency_(A)": "EUR",
        "22_Market_valuation_in_quotation_currency_(A)": "5000000",
        "23_Clean_market_valuation_in_quotation_currency_(A)": "4900000",
        "24_Market_valuation_in_portfolio_currency_(B)": "5000000",
        "25_Clean_market_valuation_in_portfolio_currency_(B)": "4900000",
        "26_Valuation_weight": "0.5",
        "32_Interest_rate_type": "Fixed",
        "33_Coupon_rate": "0.015",
        "38_Coupon_payment_frequency": "1",
        "39_Maturity_date": "2030-05-25",
        "40_Redemption_type": "Bullet",
        "41_Redemption_rate": "1",
        "46_Issuer_name": "French Republic",
        "47_Issuer_identification_code": "1A1A1A1A1A1A1A1A1A18",
        "48_Type_of_identification_code_for_issuer": "1",
        "131_Underlying_asset_category": "1",
        "1000_TPT_Version": "V7.0 (official) dated 25 November 2024",
    },
    # CIC 3: Equity
    {
        "1_Portfolio_identifying_data": "FR0010000001",
        "2_Type_of_identification_code_for_the_fund_share_or_portfolio": "1",
        "3_Portfolio_name": "Demo Bond Fund",
        "4_Portfolio_currency_(B)": "EUR",
        "5_Net_asset_valuation_of_the_portfolio_or_the_share_class_in_portfolio_currency": "10000000",
        "6_Valuation_date": "2025-12-31",
        "7_Reporting_date": "2025-12-31",
        "8_Share_price": "100",
        "8b_Total_number_of_shares": "100000",
        "11_Complete_SCR_delivery": "N",
        "12_CIC_code_of_the_instrument": "DE31",
        "14_Identification_code_of_the_instrument": "DE0007164600",  # valid ISIN (SAP)
        "15_Type_of_identification_code_for_the_instrument": "1",
        "17_Instrument_name": "SAP SE",
        "21_Quotation_currency_(A)": "EUR",
        "22_Market_valuation_in_quotation_currency_(A)": "3000000",
        "23_Clean_market_valuation_in_quotation_currency_(A)": "3000000",
        "24_Market_valuation_in_portfolio_currency_(B)": "3000000",
        "25_Clean_market_valuation_in_portfolio_currency_(B)": "3000000",
        "26_Valuation_weight": "0.3",
        "32_Interest_rate_type": "",
        "33_Coupon_rate": "",
        "38_Coupon_payment_frequency": "",
        "39_Maturity_date": "",
        "40_Redemption_type": "",
        "41_Redemption_rate": "",
        "46_Issuer_name": "SAP SE",
        "47_Issuer_identification_code": "529900D6BF99LW9R2E68",  # valid LEI
        "48_Type_of_identification_code_for_issuer": "1",
        "131_Underlying_asset_category": "3L",
        "1000_TPT_Version": "V7.0 (official) dated 25 November 2024",
    },
    # CIC 7: Cash
    {
        "1_Portfolio_identifying_data": "FR0010000001",
        "2_Type_of_identification_code_for_the_fund_share_or_portfolio": "1",
        "3_Portfolio_name": "Demo Bond Fund",
        "4_Portfolio_currency_(B)": "EUR",
        "5_Net_asset_valuation_of_the_portfolio_or_the_share_class_in_portfolio_currency": "10000000",
        "6_Valuation_date": "2025-12-31",
        "7_Reporting_date": "2025-12-31",
        "8_Share_price": "100",
        "8b_Total_number_of_shares": "100000",
        "11_Complete_SCR_delivery": "N",
        "12_CIC_code_of_the_instrument": "XL71",
        "14_Identification_code_of_the_instrument": "CASH-EUR-001",
        "15_Type_of_identification_code_for_the_instrument": "99",
        "17_Instrument_name": "Cash account EUR",
        "21_Quotation_currency_(A)": "EUR",
        "22_Market_valuation_in_quotation_currency_(A)": "2000000",
        "23_Clean_market_valuation_in_quotation_currency_(A)": "2000000",
        "24_Market_valuation_in_portfolio_currency_(B)": "2000000",
        "25_Clean_market_valuation_in_portfolio_currency_(B)": "2000000",
        "26_Valuation_weight": "0.2",
        "32_Interest_rate_type": "",
        "33_Coupon_rate": "",
        "38_Coupon_payment_frequency": "",
        "39_Maturity_date": "",
        "40_Redemption_type": "",
        "41_Redemption_rate": "",
        "46_Issuer_name": "Demo Custodian Bank",
        "47_Issuer_identification_code": "",
        "48_Type_of_identification_code_for_issuer": "",
        "131_Underlying_asset_category": "7",
        "1000_TPT_Version": "V7.0 (official) dated 25 November 2024",
    },
]

def write_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TPT V7"
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=c).value = h
    for r, row in enumerate(rows, start=2):
        for c, h in enumerate(HEADERS, start=1):
            ws.cell(row=r, column=c).value = row.get(h, "")
    wb.save(path)
    print(f"Wrote {path.relative_to(ROOT)}")


def write_csv(path: Path, rows: list[dict[str, str]], delimiter: str = ";") -> None:
    import csv
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, delimiter=delimiter)
        w.writerow(HEADERS)
        for row in rows:
            w.writerow([row.get(h, "") for h in HEADERS])
    print(f"Wrote {path.relative_to(ROOT)}")


def main() -> int:
    write_xlsx(OUT / "clean_v7.xlsx", ROWS_CLEAN)

    bad = [dict(r) for r in ROWS_CLEAN]
    bad[0]["21_Quotation_currency_(A)"] = "ZZZ"   # invalid currency
    bad[0]["38_Coupon_payment_frequency"] = "3"   # invalid coupon frequency
    bad[1]["6_Valuation_date"] = "31/12/2025"      # bad date format
    write_xlsx(OUT / "bad_formats.xlsx", bad)

    miss = [dict(r) for r in ROWS_CLEAN]
    miss[0]["5_Net_asset_valuation_of_the_portfolio_or_the_share_class_in_portfolio_currency"] = ""
    miss[0]["6_Valuation_date"] = ""
    miss[1]["5_Net_asset_valuation_of_the_portfolio_or_the_share_class_in_portfolio_currency"] = ""
    miss[1]["6_Valuation_date"] = ""
    miss[2]["5_Net_asset_valuation_of_the_portfolio_or_the_share_class_in_portfolio_currency"] = ""
    miss[2]["6_Valuation_date"] = ""
    write_csv(OUT / "missing_mandatory.csv", miss, delimiter=";")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
