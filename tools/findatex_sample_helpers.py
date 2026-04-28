"""Shared utilities for the EET/EMT/EPT sample-file generators.

Reads each template's bundled manifest JSON + spec XLSX from
``core/src/main/resources/spec/<tmpl>/`` and produces the (num, path, flag,
codification) tuples a generator needs. Also bundles a heuristic
``value_for(codif)`` so that "clean" fixtures auto-populate every
M-flagged field with a value that the validator's FormatRule will
accept.
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

# (templateId, versionToken) -> (manifest path, xlsx path) inside the JAR resources
_SPEC_LOCATIONS: dict[tuple[str, str], tuple[str, str]] = {
    ("EET", "V1.1.3"): (
        "core/src/main/resources/spec/eet/eet-v113-info.json",
        "core/src/main/resources/spec/eet/EET_V1_1_3_20260410.xlsx",
    ),
    ("EMT", "V4.3"): (
        "core/src/main/resources/spec/emt/emt-v43-info.json",
        "core/src/main/resources/spec/emt/EMT_V4_3_20251217.xlsx",
    ),
    ("EPT", "V2.1"): (
        "core/src/main/resources/spec/ept/ept-v21-info.json",
        "core/src/main/resources/spec/ept/EPT_V2_1_20221012.xlsx",
    ),
}


@dataclass(frozen=True)
class FieldRow:
    num: str           # numKey, e.g. "1", "27", "581"
    path: str          # e.g. "20040_Financial_Instrument_SFDR_Product_Type"
    flag: str          # "M", "C", "O", "" or "N"
    codif: str         # raw codification cell text


def load_spec(template_id: str, version: str) -> tuple[dict, list[FieldRow]]:
    key = (template_id, version)
    if key not in _SPEC_LOCATIONS:
        raise ValueError(f"unsupported template/version: {key}")
    manifest_rel, xlsx_rel = _SPEC_LOCATIONS[key]
    manifest = json.loads((ROOT / manifest_rel).read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(ROOT / xlsx_rel, data_only=True)
    sheet = wb[manifest["sheetName"]] if manifest["sheetName"] in wb.sheetnames else wb.worksheets[0]
    cols = manifest["columns"]
    out: list[FieldRow] = []
    for r in range(manifest["firstDataRow"], sheet.max_row + 1):
        num = sheet.cell(r, cols["numData"]).value
        path = sheet.cell(r, cols["path"]).value
        # Section-header rows have a NUM but no path; drop them.
        if path is None:
            continue
        flag = sheet.cell(r, cols["primaryFlag"]).value or ""
        codif = sheet.cell(r, cols["codification"]).value or ""
        out.append(FieldRow(
            num=str(num).strip(),
            path=str(path).strip(),
            flag=str(flag).strip().upper(),
            codif=str(codif),
        ))
    return manifest, out


# ---------- value heuristics ----------------------------------------------

# A bullet-list code prefix at line start: "1 -", "2 –", "10 -", "99.", "3L -", etc.
# Matches what CodificationParser.CLOSED_LIST_LINE recognizes as a digit-led entry.
_CLOSED_LIST_BULLET = re.compile(r"(?m)^\s*(\d+[a-zA-Z]?)\s*[-–.]")


# Hand-rolled: matches by case-insensitive substring, first hit wins. Order matters.
def value_for(codif: str) -> str:
    c = (codif or "").lower()

    # Version-literal fields (EET row 1, EMT row 1, EPT row 1) get their literal token.
    # The codification cell itself contains the literal version, so passthrough works.
    if c.strip() in {"v1.1.3", "v4.3"}:
        return codif.strip()
    if "v21 or v21uk" in c:
        return "V21"

    # Y/N variants — accept anything starting with "Y" + a separator
    if "y / n" in c or "y/n" in c or c.strip().startswith("y ,") or c.strip().startswith("y,"):
        return "Y"

    # ISO 8601 — the validator's FormatRule classifies any "iso 8601" cell as a
    # plain DATE (YYYY-MM-DD), even when the codification text mentions
    # "hh:mm:ss". So always return a bare date, never a datetime.
    if "iso 8601" in c:
        return "2025-12-31"

    # ISO 4217 currency
    if "iso 4217" in c:
        return "EUR"

    # Digit-led bullet lines ("1 - ISO 6166 ...\n2 - CUSIP ..." or
    # "Floating decimal.\n1.15% = 0.0115\n..."): the Java
    # CodificationParser treats ANY field with at least one such bullet as
    # CLOSED_LIST and rejects values that are not one of the parsed codes —
    # even when the prose mentions "ISO 6166" or "floating decimal". So
    # match the parser's threshold (>= 1 bullet) and return the first code.
    # This must come BEFORE the ISO-6166 / floating-decimal / "or" branches;
    # otherwise EMT cost cells (`Floating decimal.\n1.15% = 0.0115\n...`)
    # would be filled with `0.01` and fail the closed-list check whose only
    # entry is `1`.
    bullets = _CLOSED_LIST_BULLET.findall(codif or "")
    if len(bullets) >= 1:
        return bullets[0]

    # ISIN priority cell ("Use the following priority: - ISO 6166 code...") → valid ISIN
    if "iso 6166" in c or c.strip() == "isin":
        return "DE0007164600"  # SAP SE — valid Luhn

    # Closed list of small numeric codes ("0 / 6 / 8 / 9")
    if " / " in c and any(ch.isdigit() for ch in c) and "iso" not in c:
        # Pick the first numeric token in the cell.
        for tok in c.replace(",", " ").split():
            t = tok.strip(" /.")
            if t.isdigit():
                return t

    # "L / N" type single-letter closed lists
    if c.strip() == "l / n":
        return "L"

    # Floating decimal / percentage. Must come BEFORE the " or " branch so
    # cells like 'floating decimal or V or S or M or L or H' (EMT NUM=55)
    # are recognised as NUMERIC by the validator and answered with a number.
    if "floating decimal" in c or "percentage" in c:
        return "0.5"

    # Closed list described as "S or SF or U or N or UM or NM or ETC or B"
    if " or " in c and "iso" not in c:
        first = c.split(" or ")[0].strip().split()[-1]
        # Strip surrounding quotes/punctuation
        first = first.strip("\"' ,")
        if first:
            # If the first option is a short token (likely a code like "S", "ETC"),
            # return it verbatim. Long tokens mean the cell is a prose explanation
            # rather than a real closed list — fall back to a known-good code.
            return first.upper() if len(first) <= 4 else "S"

    # "1 to 4", "number [1 - 7]", "1, or 2, or 3", "number [1 - 6]"
    if "1 to" in c or "[1 -" in c or "1, or" in c:
        return "1"

    # Plain integer
    if c.strip() == "integer" or "(integer)" in c:
        return "100"

    # Holding period in years
    if "in years" in c:
        return "5"

    # Liquidity risk style: '"M", "I", "L"'
    if '"m"' in c and '"l"' in c:
        return "L"

    # Alphanum / free string
    if "alphanum" in c or "string" in c:
        return "Sample"

    # Fallback
    return "Sample"


# Note: ISIN/LEI checksum helpers (make_isin / make_lei) intentionally live in
# tools/build_examples.py rather than being re-exported here — the EET/EMT/EPT
# generators do not need synthetic checksum-valid identifiers (they use
# DE0007164600 directly via value_for).

# ---------- writers --------------------------------------------------------

def write_xlsx(path: Path, sheet_name: str, headers: list[str], rows: list[list[str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c).value = v
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {path.relative_to(ROOT)}")


def to_row_list(row: dict[str, str], header_nums: list[str]) -> list[str]:
    """Project a num→value dict into the column order defined by header_nums."""
    return [row[n] for n in header_nums]


def write_scenarios(
    out_dir: Path,
    sheet_name: str,
    header_nums: list[str],
    headers: list[str],
    scenarios: list[tuple[str, "Callable[[], list[dict[str, str]]]"]],
) -> None:
    """Run every (filename, factory) pair, projecting each factory's dict rows
    into the column order and writing one XLSX per scenario."""
    for filename, factory in scenarios:
        rows = [to_row_list(d, header_nums) for d in factory()]
        write_xlsx(out_dir / filename, sheet_name, headers, rows)


# ---------- self-check -----------------------------------------------------

if __name__ == "__main__":
    # Smoke: load each spec and print a few stats. Used as a manual sanity check.
    for tmpl, version in _SPEC_LOCATIONS:
        manifest, rows = load_spec(tmpl, version)
        m_count = sum(1 for r in rows if r.flag == "M")
        print(f"{tmpl} {version}: {len(rows)} fields, {m_count} M-flagged "
              f"(sheet={manifest['sheetName']!r})")
        sample_idx = [r for r in rows if r.flag == "M"][:3]
        for r in sample_idx:
            v = value_for(r.codif)
            print(f"  num={r.num} path={r.path}  →  {v!r}")
