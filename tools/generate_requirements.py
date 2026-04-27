#!/usr/bin/env python3
"""Generate requirements.md from the bundled TPT V7 spec xlsx.

The output mirrors every datapoint in the spec with its codification, the
M/C/O flag, CIC applicability and per-profile flags. Run from project root:

    python3 tools/generate_requirements.py
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SPEC = ROOT / "src" / "main" / "resources" / "spec" / "TPT_V7  20241125_updated.xlsx"
PIK = ROOT / "src" / "main" / "resources" / "spec" / "PIK guidelines 240913.xlsx"
OUT = ROOT / "requirements.md"

CIC_COLS = [
    ("CIC0", 12), ("CIC1", 13), ("CIC2", 14), ("CIC3", 15), ("CIC4", 16),
    ("CIC5", 17), ("CIC6", 18), ("CIC7", 19), ("CIC8", 20), ("CIC9", 21),
    ("CICA", 22), ("CICB", 23), ("CICC", 24), ("CICD", 25), ("CICE", 26),
    ("CICF", 27),
]

PROFILE_COLS = {
    "NW675": 29,
    "SST": 30,
    "IORP": 31,
    "EIOPA_PF.06.02.24_pos": 32,
    "EIOPA_PF.06.02.24_ass": 33,
    "EIOPA_PF.06.03.24_LT": 34,
    "ECB_PFE.06.02.30": 35,
}

FLAG_LABEL = {
    "M": "Mandatory",
    "C": "Conditional",
    "O": "Optional",
    "I": "Indicative",
    "N/A": "Not applicable",
}


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return ""
    return str(v).strip()


def section_or_field(num, path):
    """Return ('section', name) | ('field', None) | ('skip', None)."""
    if not num and not path:
        return "skip", None
    # A field row has a path; section headers do not.
    if path:
        return "field", None
    return "section", num


def detect_codification_kind(codif: str) -> str:
    txt = (codif or "").lower()
    if not txt:
        return "—"
    closed_list = re.findall(r"^\s*(\d+|\d+[a-z])\s*[-–.]", codif, flags=re.M)
    if closed_list:
        return f"closed list ({len(closed_list)} values)"
    if "iso 4217" in txt:
        return "ISO 4217 currency"
    if "iso 3166" in txt:
        return "ISO 3166-1 alpha-2 country"
    if "iso 8601" in txt or "yyyy-mm-dd" in txt:
        return "date (YYYY-MM-DD, ISO 8601)"
    if "nace" in txt:
        return "NACE V2.1"
    if "floating decimal" in txt or txt.startswith("number"):
        return "numeric (floating decimal)"
    m = re.search(r"alphanum(?:eric)?\s*\(?\s*(?:max\s*)?(\d+)\s*\)?", txt)
    if m:
        return f"alphanumeric (max {m.group(1)})"
    m = re.search(r"alpha\s*\(?\s*(\d+)\s*\)?", txt)
    if m:
        return f"alpha ({m.group(1)})"
    if "alphanumeric (4)" in txt or "cic code" in txt:
        return "CIC code (alphanumeric, 4)"
    if "alpha" in txt:
        return "alphabetic / text"
    return "free text / numeric"


def render_field(out, row, num, path, definition, codif, comment, mco, cics, profiles):
    out.append(f"### {num}\n")
    out.append(f"- **FunDataXML path:** `{path or '—'}`")
    out.append(f"- **Flag (Solvency II baseline):** {FLAG_LABEL.get(mco, mco or '—')}")
    out.append(f"- **Codification kind:** {detect_codification_kind(codif)}")
    out.append(f"- **CIC applicability:** {', '.join(cics) if cics else 'all CIC types'}")
    if profiles:
        out.append("- **Profile flags:**")
        for k, v in profiles.items():
            v_clean = str(v).replace("\n", " ").strip()
            out.append(f"    - `{k}`: {v_clean}")
    if definition:
        defn = definition.replace("\n", " ").strip()
        out.append(f"- **Definition:** {defn}")
    if codif:
        out.append("- **Codification (verbatim):**")
        out.append("  ```")
        for ln in codif.split("\n"):
            out.append(f"  {ln.rstrip()}")
        out.append("  ```")
    if comment:
        cmt = comment.replace("\n", " ").strip()
        out.append(f"- **Comment:** {cmt}")
    out.append(f"- **Source row in spec:** {row}\n")


def main() -> int:
    wb = openpyxl.load_workbook(SPEC, data_only=True)
    ws = wb["TPT V7.0"]

    out: list[str] = []
    out.append("# TPT V7 Validator — Requirements\n")
    out.append("Generated from `src/main/resources/spec/TPT_V7  20241125_updated.xlsx` "
               "(FinDatEx Tripartite Template V7.0, dated 2024-11-25).\n")
    out.append("This document mirrors every datapoint in the spec. It is the input "
               "for the rule engine — fields here drive the generated rules in "
               "`com.tpt.validator.spec.SpecCatalog`.\n")

    out.append("## Scope\n")
    out.append("- **Inputs:** TPT V7 produced files in flat `.xlsx` or `.csv` layout, "
               "header row using either NUM_DATA names (e.g. `12_CIC_code_of_the_instrument`) "
               "or the FunDataXML path (e.g. `Position / InstrumentCIC`).")
    out.append("- **Active regulatory profiles:**")
    out.append("  - **Solvency II baseline** — column K (`Mandatory / Conditional / Optional / Indicative / N/A`).")
    out.append("  - **IORP / EIOPA / ECB** — combined: column AE (IORP), AF/AG (EIOPA PF.06.02.24 positions/assets), AH (EIOPA PF.06.03.24 look-through), AI (ECB Addon PFE.06.02.30).")
    out.append("  - **NW 675** — column AC.")
    out.append("  - **SST (FINMA)** — column AD. Default off in the UI; opt-in for Swiss mandates.")
    out.append("- **Out of scope:** FunDataXML structured XML inputs, online ISIN/LEI lookups, PDF export.\n")

    out.append("## Interpretation of flags\n")
    out.append("| Flag | Meaning |")
    out.append("|------|---------|")
    out.append("| `M`  | Mandatory — must be present and well-formed. |")
    out.append("| `C`  | Conditional — must be present when CIC and other dependent fields require it. |")
    out.append("| `O`  | Optional — validated only for format/closed-list conformance if present. |")
    out.append("| `I`  | Indicative — informational, format checked, not graded as ERROR if missing. |")
    out.append("| `N/A`| Not applicable for that profile. |\n")

    out.append("## Cross-field rules (manually authored, derived from spec comments)\n")
    out.append("| ID | Description | Source |")
    out.append("|---|---|---|")
    out.append("| XF-01 | Field 11 (`CompleteSCRDelivery`) = `Y` ⇒ fields 97..105b mandatory | spec row 20 |")
    out.append("| XF-02 | Field 12 (CIC) governs applicability of every other field per CIC matrix | columns L..AA |")
    out.append("| XF-03 | Quotation/portfolio currency consistency (fields 4 vs 21 vs 24) | rows 33–36 |")
    out.append("| XF-04 | Σ field 26 (PositionWeight) ≈ 1 within tolerance | row 38 |")
    out.append("| XF-05 | Field 9 (CashPercentage) ≈ Σ market value of CIC `xx7x` / TotalNetAssets | rows 18, 22 |")
    out.append("| XF-06 | Field 5 (TotalNetAssets) ≈ field 8 × field 8b (NAV ≈ SharePrice × Shares) within precision tolerance | row 17 |")
    out.append("| XF-07 | Economic area (fields 13/74/87) consistent with country embedded in CIC | rows 23, 91, 104 |")
    out.append("| XF-08 | Field 38 (Coupon frequency) ∈ {0,1,2,4,12,52} | row 50 |")
    out.append("| XF-09 | Field 141 mandatory iff field 140 is filled | row 171 |")
    out.append("| XF-10 | Field 32 = Floating ⇒ fields 34..37 mandatory; Fixed ⇒ field 33 mandatory | rows 44–49 |")
    out.append("| XF-11 | Field 39 (Maturity) ≥ field 7 (Reporting date) for active bonds | rows 14–15, 51 |")
    out.append("| XF-12 | Field 7 (Reporting) ≥ field 6 (Valuation) | rows 14–15 |")
    out.append("| XF-13 | Field 146 (PIK) cases 1..4 — coupon and redemption fields must follow PIK guidelines | PIK guidelines |")
    out.append("| XF-14 | Field 67 (underlying CIC) mandatory iff main CIC ∈ {2, A, B, C, D, F} | row 84 |")
    out.append("| XF-15 | Field 1000 = `V7.0 (official) dated 25 November 2024` | row 179 |\n")

    out.append("## Scoring\n")
    out.append("Each profile P contributes the following category scores in `[0, 1]`:")
    out.append("- `mandatoryCompleteness[P] = 1 − missing(M-fields-for-P) / total(M-fields-for-P × applicable rows)`")
    out.append("- `formatConformance = 1 − format-errors / non-empty cells`")
    out.append("- `closedListConformance = 1 − closed-list errors / non-empty closed-list cells`")
    out.append("- `crossFieldConsistency = 1 − cross-field errors / cross-field rules evaluated`")
    out.append("- `profileCompleteness[P]` — combined M+C presence for P")
    out.append("- `overall = 0.4·mandatory + 0.2·format + 0.15·closed-list + 0.15·cross-field + 0.1·profile-completeness`\n")

    out.append("## Datapoints\n")
    out.append("(Auto-generated; do not edit by hand — re-run `tools/generate_requirements.py`.)\n")

    section = None
    field_count = 0
    for r in range(8, ws.max_row + 1):
        num = cell(ws, r, 1)
        path = cell(ws, r, 2)
        kind, _ = section_or_field(num, path)
        if kind == "skip":
            continue
        if kind == "section":
            section = num
            out.append(f"\n## Section: {section}\n")
            continue
        # field
        definition = cell(ws, r, 3)
        codif = cell(ws, r, 4)
        comment = cell(ws, r, 5)
        mco = cell(ws, r, 11)
        cics = [name for name, c in CIC_COLS if cell(ws, r, c)]
        profiles = {}
        for name, c in PROFILE_COLS.items():
            v = cell(ws, r, c)
            if v:
                profiles[name] = v
        render_field(out, r, num, path, definition, codif, comment, mco, cics, profiles)
        field_count += 1

    OUT.write_text("\n".join(out) + "\n", encoding="utf-8")
    print(f"Wrote {OUT.relative_to(ROOT)} ({field_count} datapoints, {len(out)} lines)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
